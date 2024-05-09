from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import *
from pathlib import Path
from diretorios import *
from utils.treeview_utils import load_images, create_button, save_dataframe_to_excel
import pandas as pd
import os
global df_registro_selecionado
df_registro_selecionado = None
import sqlite3
from PyQt6.QtSql import QSqlDatabase, QSqlTableModel

class EditDataDialog(QDialog):
    def __init__(self, icons_dir, model, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Editar Dados")
        self.setFixedSize(600, 400)
        self.icons_dir = Path(icons_dir)
        self.image_cache = load_images(self.icons_dir, ["plus.png", "delete.png"])
        self.model = model 
        self.setup_ui()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)

        # Criar um modelo de tabela para interagir com a aba controle_om do banco de dados
        self.model = QSqlTableModel(self)
        self.model.setTable("controle_om")
        self.model.select()  # Carregar os dados

        # Configurar para ajustar as colunas automaticamente ao conteúdo
        self.model.setEditStrategy(QSqlTableModel.EditStrategy.OnFieldChange)

        # Criar uma visualização de tabela para mostrar os dados
        self.table_view = QTableView()  # Defina table_view como um atributo de instância
        self.table_view.setModel(self.model)

        self.table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_view.setEditTriggers(QTableView.EditTrigger.DoubleClicked)
        self.table_view.verticalHeader().setVisible(False)
        # Definir o tamanho da fonte do cabeçalho da tabela
        header = self.table_view.horizontalHeader()
        font = header.font()
        font.setPointSize(12)
        header.setFont(font)

        self.setObjectName("EditDataDialog")

        self.setStyleSheet("""
            #EditDataDialog {
                background-color: black;
                color: white;
                font-size: 12pt;
                }
                QTableView {
                    border: 1px solid #d3d3d3;
                    gridline-color: #d3d3d3;
                    background-color: #f0f0f0;
                }
                QTableView::item:selected {
                    background-color: #a8a8a8;
                    color: white;
                }
                QTableView::item:hover {
                    background-color: #f5f5f5;
                    color: black;
                }
                QHeaderView::section {
                    background-color: #e0e0e0;
                    padding: 4px;
                    border: 1px solid #d3d3d3;
                    font-size: 12pt;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    border: none;
                    background: none;
                }
            """)
        # Adicionar a visualização da tabela ao layout principal
        main_layout.addWidget(self.table_view)
        # Adicionar botões para adicionar e excluir OM
        self.setupButtonsLayout()  # Chamada para o novo método
        main_layout.addLayout(self.buttons_layout)  # Adiciona o layout dos botões ao layout principal

        # Ajustar a largura das colunas para o tamanho dos valores das células
        self.table_view.resizeColumnsToContents()

    def setupButtonsLayout(self):
        self.buttons_layout = QHBoxLayout()
        self.createButtons()

    def createButtons(self):
        icon_size = QSize(40, 40)  # Tamanho do ícone para todos os botões
        self.button_specs = [
            ("Adicionar OM", self.image_cache['plus'], self.add_om, "Adicionar uma nova OM", icon_size),
            ("Excluir OM", self.image_cache['delete'], self.delete_om, "Excluir a OM selecionada", icon_size),
        ]

        for text, icon, callback, tooltip, icon_size in self.button_specs:
            btn = create_button(text=text, icon=icon, callback=callback, tooltip_text=tooltip, parent=self, icon_size=icon_size)
            self.buttons_layout.addWidget(btn)

    def add_om(self):
        # Lógica para adicionar uma nova OM
        # Adicione uma nova linha ao modelo
        self.model.insertRows(self.model.rowCount(), 1)
        # Obtenha o índice da nova linha
        new_row_index = self.model.rowCount() - 1
        # Selecione a nova linha na visualização da tabela
        self.table_view.selectRow(new_row_index)

    def delete_om(self):
        # Obter o índice da linha selecionada
        selected_indexes = self.table_view.selectionModel().selectedIndexes()

        # Se nenhum índice estiver selecionado, exibir uma mensagem de aviso
        if not selected_indexes:
            QMessageBox.warning(self, "Atenção", "Selecione uma linha para excluir.")
            return

        # Obter o índice da linha a ser excluída
        row_index = selected_indexes[0].row()

        # Remover a linha do modelo
        self.model.removeRow(row_index)

        # Confirmar as alterações no modelo
        self.model.submitAll()

        # Selecionar novamente a primeira linha para atualizar a visualização
        self.model.select()

class EditAgentesResponsaveis(QDialog):
    def __init__(self, icons_dir, model, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Editar Agentes Responsáveis")
        self.setFixedSize(800, 400)
        self.icons_dir = Path(icons_dir)
        self.image_cache = load_images(self.icons_dir, ["plus.png", "delete.png"])
        self.model = model 
        self.setup_ui()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)

        # Criar um modelo de tabela para interagir com a aba controle_om do banco de dados
        self.model = QSqlTableModel(self)
        self.model.setTable("controle_agentes_responsaveis")
        self.model.select()  # Carregar os dados

        # Configurar para ajustar as colunas automaticamente ao conteúdo
        self.model.setEditStrategy(QSqlTableModel.EditStrategy.OnFieldChange)

        # Criar uma visualização de tabela para mostrar os dados
        self.table_view = QTableView()
        self.table_view.setModel(self.model)
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        self.table_view.setEditTriggers(QTableView.EditTrigger.DoubleClicked)

        self.table_view.verticalHeader().setVisible(False)

        # Definir o tamanho da fonte do cabeçalho da tabela
        header = self.table_view.horizontalHeader()
        font = header.font()
        font.setPointSize(12)
        header.setFont(font)
    
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)  # Ajustar todas as colunas para preencher o espaço

        self.setObjectName("EditDataDialog")

        self.setStyleSheet("""
            #EditDataDialog {
                background-color: black;
                color: white;
                font-size: 12pt;
                }
                QTableView {
                    border: 1px solid #d3d3d3;
                    gridline-color: #d3d3d3;
                    background-color: #f0f0f0;
                }
                QTableView::item:selected {
                    background-color: #a8a8a8;
                    color: white;
                }
                QTableView::item:hover {
                    background-color: #f5f5f5;
                    color: black;
                }
                QHeaderView::section {
                    background-color: #e0e0e0;
                    padding: 4px;
                    border: 1px solid #d3d3d3;
                    font-size: 12pt;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    border: none;
                    background: none;
                }
            """)
        self.delegate = ComboBoxDelegate(self.table_view)
        self.table_view.setItemDelegateForColumn(self.model.fieldIndex("funcao"), self.delegate)
        
        main_layout.addWidget(self.table_view)
        # Adicionar botões para adicionar e excluir OM
        self.setupButtonsLayout()  # Chamada para o novo método
        main_layout.addLayout(self.buttons_layout)  # Adiciona o layout dos botões ao layout principal

    def setupButtonsLayout(self):
        self.buttons_layout = QHBoxLayout()
        self.createButtons()

    def createButtons(self):
        icon_size = QSize(40, 40)  # Tamanho do ícone para todos os botões
        self.button_specs = [
            ("Adicionar Responsável", self.image_cache['plus'], self.add_responsavel, "Adicionar novo Agente Responsável", icon_size),
            ("Excluir Responsável", self.image_cache['delete'], self.delete_responsavel, "Excluir Agente Responsável", icon_size),
        ]

        for text, icon, callback, tooltip, icon_size in self.button_specs:
            btn = create_button(text=text, icon=icon, callback=callback, tooltip_text=tooltip, parent=self, icon_size=icon_size)
            self.buttons_layout.addWidget(btn)

    def add_responsavel(self):
        row_count = self.model.rowCount()
        self.model.insertRow(row_count)
        self.table_view.selectRow(row_count)
        # Defina valores padrão para a nova linha aqui, se necessário
        # Por exemplo, self.model.setData(self.model.index(row_count, 1), "Valor Padrão")
        self.model.submitAll()  # Confirme a adição imediatamente

    def delete_responsavel(self):
        # Obter o índice da linha selecionada
        selected_indexes = self.table_view.selectionModel().selectedIndexes()

        # Se nenhum índice estiver selecionado, exibir uma mensagem de aviso
        if not selected_indexes:
            QMessageBox.warning(self, "Atenção", "Selecione uma linha para excluir.")
            return

        # Obter o índice da linha a ser excluída
        row_index = selected_indexes[0].row()

        # Remover a linha do modelo
        self.model.removeRow(row_index)

        # Confirmar as alterações no modelo
        self.model.submitAll()

        # Selecionar novamente a primeira linha para atualizar a visualização
        self.model.select()

class ComboBoxDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.items = [
            "Ordenador de Despesas",
            "Ordenador de Despesas Substituto",
            "Agente Fiscal",
            "Encarregado da Divisão de Obtenção",
            "Ajudante da Divisão de Obtenção",
            "Supervisor da Seção de Licitação",
            "Auxiliar da Seção de Licitação",
            "Assessor Jurídico",
            "Agente Financeiro",
            "Agente Financeiro Substituto"
        ]

    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        editor.addItems(self.items)
        editor.installEventFilter(self)

        # Conectar a mudança de índice para submeter as alterações
        editor.currentIndexChanged.connect(lambda: self.commitAndCloseEditor(editor))
        
        return editor

    def commitAndCloseEditor(self, editor):
        self.commitData.emit(editor)
        self.closeEditor.emit(editor)

    def setEditorData(self, editor, index):
        text = index.data(Qt.ItemDataRole.DisplayRole)
        idx = editor.findText(text)
        if idx >= 0:
            editor.setCurrentIndex(idx)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)
        # Submeter todas as alterações ao modelo assim que o usuário selecionar um novo item
        model.submitAll()

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)

class SettingsDialog(QDialog):
    def __init__(self, config_manager, parent=None):
        super().__init__(parent)
        self.config_manager = config_manager
        self.database_path = Path(CONTROLE_DADOS)
        self.config_file = CONFIG_FILE
        self.pasta_base = Path(self.load_config('save_location', str(Path.home() / 'Desktop')))
        self.setWindowTitle("Configurações")
        self.setFixedSize(600, 400)  # Tamanho total da janela de diálogo
        self.parent_app = parent
        self.model = QSqlTableModel(self)  # Criar o modelo
        self.setup_ui()

    def setup_ui(self):
        self.setStyleSheet("""
        QDialog {
            font-size: 12pt;
            color: #333;
            background-color: #f0f0f0;
        }
        QGroupBox {
            font-size: 12pt;
            border: 2px solid #6c6c6c;
            border-radius: 5px;
            margin-top: 2ex;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 3px 0 3px;
            color: #444;
        }
        QLineEdit, QPushButton, QLabel, QComboBox, QTableView{
            font-size: 12pt;
        }
        """)
        main_layout = QVBoxLayout(self)  # Único layout vertical para todo o diálogo

        # Configurar OM/UASG
        om_uasg_layout = QVBoxLayout()  
        om_uasg_layout.addWidget(QLabel("Configurar OM/UASG"))

        # Adicionando os três botões em um layout horizontal
        buttons_layout = QHBoxLayout()
        self.gerar_tabela_btn = QPushButton("Gerar Tabela")
        self.importar_tabela_btn = QPushButton("Importar Tabela")
        self.editar_dados_btn = QPushButton("Editar Dados")
        self.gerar_tabela_btn.clicked.connect(self.generate_table)
        self.importar_tabela_btn.clicked.connect(self.update_om)
        self.editar_dados_btn.clicked.connect(self.edit_data)
        buttons_layout.addWidget(self.gerar_tabela_btn)
        buttons_layout.addWidget(self.importar_tabela_btn)
        buttons_layout.addWidget(self.editar_dados_btn)

        # Adicionando o layout dos botões ao layout principal
        om_uasg_layout.addLayout(buttons_layout)
        main_layout.addLayout(om_uasg_layout)

        # Configurar Local de Salvamento dos Arquivos
        file_save_layout = QHBoxLayout()
        file_save_btn = QPushButton("Definir novo local")
        file_save_btn.clicked.connect(self.define_file_save_location)
        file_save_layout.addWidget(QLabel("Local de Salvamento dos Arquivos"))
        file_save_layout.addWidget(file_save_btn)
        main_layout.addLayout(file_save_layout)

        # Configurar Agentes Responsáveis
        agentes_responsaveis_layout = QVBoxLayout()
        agentes_responsaveis_layout.addWidget(QLabel("Definir Agentes Responsáveis"))
        
        button_responsaveis_layout = QHBoxLayout()
        self.gerar_tabela_responsaveis_btn = QPushButton("Gerar Tabela")
        self.importar_tabela_responsaveis_btn = QPushButton("Importar Tabela")
        self.editar_dados_responsaveis_btn = QPushButton("Editar Dados")
        self.gerar_tabela_responsaveis_btn.clicked.connect(self.generate_table_agentes_responsaveis)
        self.importar_tabela_responsaveis_btn.clicked.connect(self.update_agentes_responsaveis)
        self.editar_dados_responsaveis_btn.clicked.connect(self.edit_data_agentes_responsaveis)
        button_responsaveis_layout.addWidget(self.gerar_tabela_responsaveis_btn)
        button_responsaveis_layout.addWidget(self.importar_tabela_responsaveis_btn)
        button_responsaveis_layout.addWidget(self.editar_dados_responsaveis_btn)

        # Adicionando o layout dos botões ao layout principal
        agentes_responsaveis_layout.addLayout(button_responsaveis_layout)
        main_layout.addLayout(agentes_responsaveis_layout)

        # Configurar carregamento de tabela
        carregar_tabela_layout = QHBoxLayout()
        carregar_tabela_btn = QPushButton("Carregar Tabela")
        carregar_tabela_btn.clicked.connect(self.safe_load_table)
        carregar_tabela_layout.addWidget(QLabel("Carregar tabela excel ou libre"))
        carregar_tabela_layout.addWidget(carregar_tabela_btn)
        main_layout.addLayout(carregar_tabela_layout)

        # Configurar atualização de banco de dados
        carregar_database_layout = QHBoxLayout()
        carregar_database_btn = QPushButton("Atualizar Banco de Dados")
        carregar_database_btn.clicked.connect(self.safe_update_database)
        carregar_database_layout.addWidget(QLabel("Carregar dados de um arquivo .db"))
        carregar_database_layout.addWidget(carregar_database_btn)
        main_layout.addLayout(carregar_database_layout)

    def edit_data_agentes_responsaveis(self):
        # Conectar ao banco de dados
        with sqlite3.connect(self.database_path) as conn:
            cursor = conn.cursor()

            # Verificar se a tabela existe e criá-la se não existir
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS controle_agentes_responsaveis (
                    nome TEXT,
                    funcao TEXT,
                    posto TEXT
                );
            """)

            # Fechar o cursor
            cursor.close()

        # Abrir o QDialog para editar os dados, passando o modelo como argumento
        dialog = EditAgentesResponsaveis(ICONS_DIR, self.model, self)
        dialog.exec()

        # Atualizar o modelo de tabela após fechar o QDialog para refletir quaisquer alterações
        self.model.select()

    def load_config(self, key, default_value):
        try:
            with open(self.config_file, 'r') as f:
                config = json.load(f)
                return config.get(key, default_value)
        except (FileNotFoundError, json.JSONDecodeError):
            return default_value

    def save_config(self, key, value):
        config = {}
        try:
            with open(self.config_file, 'r') as f:
                config = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            pass
        config[key] = value
        with open(self.config_file, 'w') as f:
            json.dump(config, f)

    def define_file_save_location(self):
        file_path = QFileDialog.getExistingDirectory(self, "Selecionar Pasta")
        if file_path:
            self.pasta_base = Path(file_path)
            self.config_manager.update_config('save_location', str(self.pasta_base))
            print(f"Local de salvamento definido: {self.pasta_base}")

    def edit_data(self):
        # Abrir o QDialog para editar os dados, passando o modelo como argumento
        dialog = EditDataDialog(ICONS_DIR, self.model, self)
        dialog.exec()

        # Atualizar o modelo de tabela após fechar o QDialog para refletir quaisquer alterações
        self.model.select()

    def safe_load_table(self):
        try:
            self.parent_app.load_table()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar a tabela: {str(e)}")

    def safe_update_database(self):
        try:
            self.parent_app.update_database()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar o banco de dados: {str(e)}")

    def generate_table_agentes_responsaveis(self):
        # Criar DataFrame com os dados
        data = {
            'nome': ['Ex: NOME'],  
            'funcao': ['Função exemplo: Ordenador de Despesas'],
            'posto': ['Posto exemplo: Capitão de Fragata (IM)']
        }
        df = pd.DataFrame(data)

        # Adicionar exemplos aos cabeçalhos das colunas
        df.columns = ['nome', 'funcao', 'posto']

        # Adicionar comentário indicando a primeira linha como exemplo
        df.to_excel(
            "tabela_agentes_responsaveis.xlsx",
            index=False,
            engine='openpyxl',
        )

        # Definir largura das colunas
        from openpyxl import load_workbook
        wb = load_workbook("tabela_agentes_responsaveis.xlsx")
        ws = wb.active
        ws.column_dimensions['A'].width = 100  # Coluna 'nome'
        ws.column_dimensions['B'].width = 50  # Coluna 'funcao'
        ws.column_dimensions['C'].width = 50  # Coluna 'posto'
        wb.save("tabela_agentes_responsaveis.xlsx")

        file_path = "tabela_agentes_responsaveis.xlsx"  # Defina o caminho do arquivo aqui

        # Abrir o arquivo xlsx após ser criado
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(file_path)))

    def update_agentes_responsaveis(self):
        # Supondo que import_uasg_to_db atualize o banco de dados corretamente
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Selecione o arquivo Excel",
            "",  
            "Arquivos Excel (*.xlsx);;Todos os Arquivos (*)"
        )

        if filename:
            self.import_agentes_to_db(filename)

    def import_agentes_to_db(self, filepath):
        # Ler os dados do arquivo Excel
        df = pd.read_excel(filepath, usecols=['nome', 'funcao', 'posto'])
        
        # Conectar ao banco de dados
        with sqlite3.connect(self.database_path) as conn:
            cursor = conn.cursor()
            
            # Verificar se a tabela existe
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS controle_agentes_responsaveis (
                    nome TEXT,
                    funcao TEXT,
                    posto TEXT
                );
            """)
            
            # Se a tabela já existe ou foi criada, adiciona os dados
            df.to_sql('controle_agentes_responsaveis', conn, if_exists='replace', index=False)
            
            # Fechar o cursor
            cursor.close()

    def generate_table(self):
        # Criar DataFrame com os dados
        data = {
            'uasg': ['Ex: 787010'],  
            'orgao_responsavel': ['Ex: CENTRO DE INTENDÊNCIA DA MARINHA EM BRASÍLIA'],
            'sigla_om': ['Ex: CeIMBra']
        }
        df = pd.DataFrame(data)

        # Adicionar exemplos aos cabeçalhos das colunas
        df.columns = ['uasg', 'orgao_responsavel', 'sigla_om']

        # Adicionar comentário indicando a primeira linha como exemplo
        df.to_excel(
            "tabela_uasg.xlsx",
            index=False,
            engine='openpyxl',
        )

        # Definir largura das colunas
        from openpyxl import load_workbook
        wb = load_workbook("tabela_uasg.xlsx")
        ws = wb.active
        ws.column_dimensions['A'].width = 15  # Coluna 'uasg'
        ws.column_dimensions['B'].width = 60  # Coluna 'orgao_responsavel'
        ws.column_dimensions['C'].width = 15  # Coluna 'sigla_om'
        wb.save("tabela_uasg.xlsx")

        file_path = "tabela_uasg.xlsx"  # Defina o caminho do arquivo aqui

        # Abrir o arquivo xlsx após ser criado
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(file_path)))

    def update_om(self):
        # Supondo que import_uasg_to_db atualize o banco de dados corretamente
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Selecione o arquivo Excel",
            "",  
            "Arquivos Excel (*.xlsx);;Todos os Arquivos (*)"
        )

        if filename:
            self.import_uasg_to_db(filename)

    def import_uasg_to_db(self, filepath):
        # Ler os dados do arquivo Excel
        df = pd.read_excel(filepath, usecols=['uasg', 'orgao_responsavel', 'sigla_om'])
        
        # Conectar ao banco de dados e criar a tabela se não existir
        with sqlite3.connect(self.database_path) as conn:
            df.to_sql('controle_om', conn, if_exists='replace', index=False)  # Use 'replace' para substituir ou 'append' para adicionar
