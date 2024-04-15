from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import *
from diretorios import *
import pandas as pd
import os
global df_registro_selecionado
df_registro_selecionado = None
import subprocess
from docxtpl import DocxTemplate
PLANEJAMENTO_DIR = BASE_DIR / "planejamento"
import sys
import shutil
import tempfile
import os
from openpyxl import load_workbook

NOME_COLUNAS = {
    'mod': 'Mod.',
    'num_pregao': 'N',
    'ano_pregao': 'Ano',
    'item_pca': 'Item PCA',
    'portaria_PCA': 'Portaria_PCA',	
    'data_sessao': 'Data Sessão',
    'nup': 'NUP',
    'objeto': 'Objeto',
    'uasg': 'UASG',
    'orgao_responsavel': 'Órgão Responsável',
    'sigla_om': 'Sigla Órgão',
    'setor_responsavel': 'Demandante',
    'coordenador_planejamento': 'Coordenador',
    'etapa': 'Etapa',
    'pregoeiro': 'Pregoeiro',
}

class AutorizacaoAberturaLicitacaoDialog(QDialog):
    def __init__(self, main_app, df_registro, parent=None):
        super().__init__(parent)
        self.main_app = main_app
        self.df_registro = df_registro

        self.item_pca = df_registro['item_pca'].iloc[0]
        self.portaria_PCA = df_registro['portaria_PCA'].iloc[0]
        self.id_processo = df_registro['id_processo'].iloc[0]
        
        self.setWindowTitle("Autorização para Abertura")
        self.setFixedSize(300, 400)
        self.layout = QVBoxLayout(self)
        self.pasta = ''
        self.setupUi()

    def setupUi(self):
        settings = QSettings("SuaOrganizacao", "SeuAplicativo")
        # Grupo 1: Autoridade Competente
        self.grupoAutoridade = QGroupBox("Autoridade Competente")
        self.grupoAutoridadeLayout = QVBoxLayout(self.grupoAutoridade)
        self.ordenadordespesasComboBox = QComboBox()
        self.carregarOrdenadorDespesas()
        self.grupoAutoridadeLayout.addWidget(self.ordenadordespesasComboBox)
        self.layout.addWidget(self.grupoAutoridade)
        
        # Grupo 2: Seleção de Pasta
        self.grupoSelecaoPasta = QGroupBox("Local de Salvamento do Arquivo")
        self.grupoSelecaoPastaLayout = QVBoxLayout(self.grupoSelecaoPasta)
        self.labelPasta = QLabel("Selecionar pasta para salvar o arquivo:")
        iconPathFolder = ICONS_DIR / "abrir_pasta.png"

        self.botaoSelecionarPasta = QPushButton("  Selecionar Pasta")
        self.botaoSelecionarPasta.setIcon(QIcon(str(iconPathFolder)))  # Converter Path para string
        self.botaoSelecionarPasta.clicked.connect(self.selecionarPasta)
        self.grupoSelecaoPastaLayout.addWidget(self.labelPasta)
        self.grupoSelecaoPastaLayout.addWidget(self.botaoSelecionarPasta)
        self.layout.addWidget(self.grupoSelecaoPasta)

        # Grupo 3: Edição do Template
        self.grupoEdicaoTemplate = QGroupBox("Edição do Modelo")
        self.grupoEdicaoTemplateLayout = QVBoxLayout(self.grupoEdicaoTemplate)
        self.labelEdicao = QLabel("Editar o arquivo modelo de Autorização:")
        iconPathEdit = ICONS_DIR / "text.png"
        
        self.botaoEdicaoTemplate = QPushButton("  Editar Modelo")
        self.botaoEdicaoTemplate.setIcon(QIcon(str(iconPathEdit)))  # Converter Path para string
        self.botaoEdicaoTemplate.clicked.connect(self.editarTemplate)
        self.grupoEdicaoTemplateLayout.addWidget(self.labelEdicao)
        self.grupoEdicaoTemplateLayout.addWidget(self.botaoEdicaoTemplate)
        self.layout.addWidget(self.grupoEdicaoTemplate)

        # Grupo 4: Edição do Template - PCA
        self.grupoItemPCA = QGroupBox("Plano de Contratações Anual (PCA)")
        self.grupoItemPCALayout = QVBoxLayout(self.grupoItemPCA)

        # Cria um layout horizontal para o item do PCA
        self.itemPCALayout = QHBoxLayout()
        self.labelItemPCA = QLabel("Item do PCA:")
        self.lineEditItemPCA = QLineEdit()
        self.lineEditItemPCA.setText(str(self.item_pca))
        # Adiciona o label e o line edit ao layout horizontal
        self.itemPCALayout.addWidget(self.labelItemPCA)
        self.itemPCALayout.addWidget(self.lineEditItemPCA)
        self.grupoItemPCALayout.addLayout(self.itemPCALayout)
        self.lineEditItemPCA.editingFinished.connect(self.salvarItemPCA)

        # Continuação para adicionar a Portaria e seu QLineEdit
        self.labelPortariaPCA = QLabel("Portaria:")
        self.lineEditPortariaPCA = QLineEdit()
        self.lineEditPortariaPCA.setText(str(self.portaria_PCA))
        # Carregar valor pré-definido da portaria de QSettings
        portaria_predefinida = settings.value("portaria_PCA", f"{self.portaria_PCA}")
        self.lineEditPortariaPCA.setPlaceholderText("Ex: 05 Ceimbra, de 26 de janeiro de 2024.")
        # Conectar o sinal de edição concluída do QLineEdit da portaria a uma função para salvar o valor

        self.lineEditPortariaPCA.editingFinished.connect(self.salvarPortariaPCA)

        self.grupoItemPCALayout.addWidget(self.labelPortariaPCA)
        self.grupoItemPCALayout.addWidget(self.lineEditPortariaPCA)
        self.layout.addWidget(self.grupoItemPCA)

        # Grupo 5: Criação de Documento
        self.grupoCriacaoDocumento = QGroupBox("Criação de Documento")
        self.grupoCriacaoDocumentoLayout = QVBoxLayout(self.grupoCriacaoDocumento)
        self.botoesLayout = QHBoxLayout()
        # Caminhos dos ícones
        iconPathDocx = ICONS_DIR / "word.png"
        iconPathPdf = ICONS_DIR / "pdf64.png"
        # Botões com ícones
        self.botaoDocx = QPushButton("  Docx")
        self.botaoDocx.setIcon(QIcon(str(iconPathDocx)))  # Converter Path para string
        self.botaoPdf = QPushButton("  Pdf")
        self.botaoPdf.setIcon(QIcon(str(iconPathPdf)))  # Converter Path para string
        
        self.botaoDocx.clicked.connect(self.gerarDocx)
        
        self.botaoPdf.clicked.connect(self.gerarPdf)
        self.botoesLayout.addWidget(self.botaoDocx)
        self.botoesLayout.addWidget(self.botaoPdf)
        self.grupoCriacaoDocumentoLayout.addLayout(self.botoesLayout)
        self.layout.addWidget(self.grupoCriacaoDocumento)

        # Aplicar estilo de borda aos grupos
        estiloBorda = "QGroupBox { border: 1px solid gray; border-radius: 5px; margin-top: 0.5em; } " \
                      "QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px 0 3px; }"
        self.grupoAutoridade.setStyleSheet(estiloBorda)
        self.grupoSelecaoPasta.setStyleSheet(estiloBorda)
        self.grupoEdicaoTemplate.setStyleSheet(estiloBorda)
        self.grupoItemPCA.setStyleSheet(estiloBorda)
        self.grupoCriacaoDocumento.setStyleSheet(estiloBorda)

    def salvarPortariaPCA(self):
        # Atualiza o valor de portaria_PCA em QSettings
        self.portaria_PCA = self.lineEditPortariaPCA.text()
        self.salvarAlteracoesExcel()

    def salvarItemPCA(self):
        # Atualiza o valor de portaria_PCA em QSettings
        self.item_pca = self.lineEditItemPCA.text()
        self.salvarAlteracoesExcel()

        
    def salvarAlteracoesExcel(self):
        settings = QSettings("SuaOrganizacao", "SeuAplicativo")
        settings.setValue("portaria_PCA", self.portaria_PCA)
        settings.setValue("item_pca", self.item_pca)

        try:
            # Carrega a planilha Excel
            workbook = load_workbook(filename=CONTROLE_PROCESSOS_DIR)
            sheet = workbook.active

            for row in range(2, sheet.max_row + 1):
                if (sheet[f'A{row}'].value == self.mod and
                    sheet[f'B{row}'].value == int(self.num_pregao) and
                    sheet[f'C{row}'].value == int(self.ano_pregao)):

                    # Atualiza as células no Excel com os novos valores
                    sheet[f'D{row}'].value = self.item_pca
                    sheet[f'E{row}'].value = self.portaria_PCA
                    break

            # Salva o arquivo Excel
            workbook.save(filename=CONTROLE_PROCESSOS_DIR)
            QMessageBox.information(self, "Sucesso", "As alterações foram salvas com sucesso.")

        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Não foi possível salvar as alterações: {e}")

        # Depois de salvar, chama a função para recarregar e atualizar o QTreeView
        self.atualizarTreeView()

    def atualizarTreeView(self):
        # Recarrega os dados do DataFrame a partir do arquivo Excel atualizado
        self.main_app.df_licitacao_completo = pd.read_excel(CONTROLE_PROCESSOS_DIR)

        # Limpa o modelo atual
        self.main_app.model.clear()

        # Reaplica os cabeçalhos das colunas
        self.main_app.model.setHorizontalHeaderLabels([self.main_app.NOME_COLUNAS[col] for col in self.main_app.NOME_COLUNAS])

        # Repopula o modelo com os dados atualizados
        for _, row in self.main_app.df_licitacao_completo.iterrows():
            items = [QStandardItem(str(row[col])) for col in self.main_app.columns_treeview]
            self.main_app.model.appendRow(items)

        # Ajusta a largura das colunas com base nos dados atualizados
        self.main_app._adjust_column_widths()

    def editarTemplate(self):
        template_path = PLANEJAMENTO_DIR / "template_autorizacao.docx"
        try:
            if sys.platform == "win32":
                subprocess.run(["start", "winword", str(template_path)], check=True, shell=True)
            elif sys.platform == "darwin":  # macOS
                subprocess.run(["open", str(template_path)], check=True)
            else:  # linux variants
                subprocess.run(["xdg-open", str(template_path)], check=True)
        except subprocess.CalledProcessError as e:
            QMessageBox.warning(self, "Erro", f"Não foi possível abrir o documento: {e}")
            
    def carregarOrdenadorDespesas(self):
        try:
            self.ordenador_despesas_df = pd.read_excel(ORDENADOR_DESPESAS_DIR)
            for index, row in self.ordenador_despesas_df.iterrows():
                texto_display = f"{row['nome']}\n{row['posto']}\n{row['od']}"
                self.ordenadordespesasComboBox.addItem(texto_display, userData=row.to_dict())
        except Exception as e:
            print(f"Erro ao carregar tabela Ordenador de Despesas: {e}")

    def selecionarPasta(self):
        self.pasta = QFileDialog.getExistingDirectory(self, "Selecionar Pasta")
        if self.pasta:
            print(f"Pasta selecionada: {self.pasta}")
    
    def gerarDocx(self):
        try:
            # Verifica se um registro foi selecionado na tabela
            if self.df_registro is None:
                QMessageBox.warning(None, "Seleção Necessária", "Por favor, selecione um registro na tabela antes de gerar um documento.")
                return

            # Continua com o processo de geração do documento
            template_path = PLANEJAMENTO_DIR / "template_autorizacao.docx"
            id_processo_formatado = self.df_registro['id_processo'].iloc[0].replace('/', '-')
            salvar_nome = f"{id_processo_formatado} - Autorizacao para abertura de Processo Administrativo"

            temp_dir = tempfile.mkdtemp()
            temp_template_path = shutil.copy(template_path, temp_dir)
            doc = DocxTemplate(temp_template_path)

            nome_selecionado = self.ordenadordespesasComboBox.currentText()
            valor_completo = self.ordenadordespesasComboBox.currentData(Qt.ItemDataRole.UserRole)

            # Preparar os dados para renderizar no template
            data = {
                **self.df_registro.to_dict(orient='records')[0],  # Incorporar dados do DataFrame
                "item_pca": self.item_pca,  # Usar o valor de item_pca passado como argumento
                "portaria_PCA": self.portaria_PCA,  # Usar o valor de portaria_PCA passado como argumento
                'ordenador_de_despesas': f"{valor_completo['nome']}\n{valor_completo['posto']}\n{valor_completo['od']}"  # Utilizar a string formatada
            }

            # Substituições adicionais conforme especificado
            id_processo = self.df_registro['id_processo'].iloc[0]
            if "PE" in id_processo:
                pregao_num = id_processo.split()[1]
                id_processo = f"Pregão Eletrônico nº {pregao_num}"
            elif "CC" in id_processo:
                concorrencia_num = id_processo.split()[1]
                id_processo_substituido = f"Concorrência nº {concorrencia_num}"
            else:
                id_processo_substituido = id_processo

            data["id_processo"] = id_processo_substituido

            doc.render(data)
            save_path = os.path.join(self.pasta, f"{salvar_nome}.docx")
            doc.save(save_path)
            shutil.rmtree(temp_dir)  # Limpar a cópia temporária ao concluir

            QMessageBox.information(None, "Sucesso", "Documento DOCX gerado com sucesso no diretório selecionado.")
            os.startfile(save_path)

        except Exception as e:
            QMessageBox.warning(None, "Erro", f"Erro ao gerar documento DOCX: {e}")



    def gerarPdf(self):
        # Lógica para gerar documento PDF
        pass
