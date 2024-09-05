from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import *
from modules.planejamento.utilidades_planejamento import DatabaseManager
from modules.dispensa_eletronica.documentos_cp_dfd_tr import PDFAddDialog, ConsolidarDocumentos, load_config_path_id
from diretorios import *
from pathlib import Path
import pandas as pd
import os
import subprocess
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sqlite3
from fpdf import FPDF 
import webbrowser

class RealLineEdit(QLineEdit):
    def __init__(self, text='', parent=None):
        super().__init__(text, parent)
        self.setText(self.format_to_real(self.text()))

    def focusInEvent(self, event):
        # Remove the currency formatting when the user focuses on the widget
        self.setText(self.format_to_plain_number(self.text()))
        super().focusInEvent(event)
    
    def focusOutEvent(self, event):
        # Add the currency formatting when the user leaves the widget
        self.setText(self.format_to_real(self.text()))
        super().focusOutEvent(event)
    
    def format_to_real(self, value):
        try:
            # Convert the plain number to real currency format
            value = float(value.replace('.', '').replace(',', '.').strip())
            return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except ValueError:
            return value
    
    def format_to_plain_number(self, value):
        try:
            # Convert the real currency format to plain number
            value = float(value.replace('R$', '').replace('.', '').replace(',', '.').strip())
            return f"{value:.2f}".replace('.', ',')
        except ValueError:
            return value

class EditDataDialog(QDialog):
    dados_atualizados = pyqtSignal()
    title_updated = pyqtSignal(str)

    def __init__(self, df_registro_selecionado, icons_dir, parent=None):
        super().__init__(parent)
        self.df_registro_selecionado = df_registro_selecionado
        self.ICONS_DIR = Path(icons_dir)

        self.navigation_buttons = []
        self.consolidador = ConsolidarDocumentos(df_registro_selecionado)
        self._init_paths()

        # Inicializar self.formulario_excel antes de _init_ui
        self.formulario_excel = FormularioExcel(self.df_registro_selecionado, self.pasta_base, self)

        self._init_ui()
        self._init_connections()

    def _init_paths(self):
        self.database_path = Path(load_config("CONTROLE_DADOS", str(CONTROLE_DADOS)))
        self.database_manager = DatabaseManager(self.database_path)
        self.config = load_config_path_id()
        self.pasta_base = Path(self.config.get('pasta_base', str(Path.home() / 'Desktop')))

    def _init_ui(self):
        # Define o título da janela e o ícone
        self.setWindowTitle("Editar Dados do Processo")
        
        # Define o ícone da janela como 'edit.png'
        icon_path = self.ICONS_DIR / "edit.png"
        if icon_path.is_file():
            self.setWindowIcon(QIcon(str(icon_path)))
        else:
            print(f"Icon not found: {icon_path}")
        self.setFixedSize(1200, 620)

        # Layout principal vertical para os componentes existentes
        layout_principal = QVBoxLayout()

        # Adicionando título e navegação
        layout_principal.addWidget(self.update_title_label())
        layout_principal.addLayout(self.create_navigation_layout())

        layout_principal.setSpacing(0)
        layout_principal.setContentsMargins(0, 0, 0, 0)

        # Gerenciador de Stacked Widgets e Configuração
        self.stack_manager = QStackedWidget(self)
        self.setup_stacked_widgets()
        layout_principal.addWidget(self.stack_manager)

        # Cria o layout de agentes responsáveis e aplica borda lateral
        layout_agentes_responsaveis = self.create_agentes_responsaveis_layout()

        # Layout horizontal principal para conter ambos os layouts
        hlayout_main = QHBoxLayout(self)
        hlayout_main.addLayout(layout_principal)  # Adiciona o layout principal à esquerda
        hlayout_main.addWidget(layout_agentes_responsaveis)  # Adiciona o layout de agentes à direita

        # Define o layout principal como o layout horizontal
        self.setLayout(hlayout_main)

        # Mostra o widget inicial
        self.show_widget("Informações Gerais")
    
    def _init_connections(self):
        self.title_updated.connect(self.update_title_label_text)

    def create_navigation_layout(self):
        nav_layout = QHBoxLayout()

        nav_layout.setSpacing(0)
        nav_layout.setContentsMargins(0, 0, 0, 0)
        
        # Cria um QIcon ao invés de QPixmap
        brasil_icon = QIcon(str(BRASIL_IMAGE_PATH))
        
        # Cria um QLabel e adiciona o ícone como um QIcon
        image_label_esquerda = QLabel()
        image_label_esquerda.setAlignment(Qt.AlignmentFlag.AlignCenter)
        image_label_esquerda.setPixmap(brasil_icon.pixmap(30, 30))  # Define o QIcon como um QPixmap para o QLabel
        
        nav_layout.addWidget(image_label_esquerda)

        # Lista de botões de navegação
        buttons = [
            ("Informações Gerais", "Informações Gerais"),
            ("Setor Responsável", "Setor Responsável"),
            ("Documentos", "Documentos"),
            ("Anexos", "Anexos"),
        ]

        for text, name in buttons:
            self.add_navigation_button(nav_layout, text, lambda _, n=name: self.show_widget(n))

        nav_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))
        self.add_action_buttons(nav_layout)
        nav_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))

        return nav_layout

    def show_widget(self, name):
        # Desmarcar todos os botões de navegação
        for button in self.navigation_buttons:
            button.setChecked(False)
        # Encontrar o botão correspondente e marcar
        for button in self.navigation_buttons:
            if button.text() == name:
                button.setChecked(True)
                self.update_button_styles(button)
                break
        # Mostrar o widget correspondente no QStackedWidget
        for i in range(self.stack_manager.count()):
            widget = self.stack_manager.widget(i)
            if widget.objectName() == name:
                self.stack_manager.setCurrentWidget(widget)
                break

    def add_navigation_button(self, layout, text, callback):
        button = QPushButton(text)
        button.setCheckable(True)
        button.setMinimumWidth(172)
        button.setStyleSheet(self.get_button_style())
        button.clicked.connect(callback)
        layout.addWidget(button)
        self.navigation_buttons.append(button)

    def get_button_style(self):
        return (
            "QPushButton {"
            "border: 1px solid #414242; background: #B0B0B0; color: black; font-weight: bold; font-size: 12pt;"
            "border-top-left-radius: 5px; border-top-right-radius: 5px; "
            "border-bottom-left-radius: 0px; border-bottom-right-radius: 0px; "
            "border-bottom-color: #414242; }"
            "QPushButton:hover { background: #D0D0D0; font-weight: bold; color: black; }"
        )

    def update_button_styles(self, active_button):
        for button in self.navigation_buttons:
            if button == active_button:
                button.setStyleSheet(
                    "QPushButton { border: 1px solid #414242; background: #414242; font-weight: bold; color: white; "
                    "border-top-left-radius: 5px; border-top-right-radius: 5px; "
                    "border-bottom-left-radius: 0px; border-bottom-right-radius: 0px; "
                    "border-bottom-color: #414242; font-size: 12pt; }"
                    "QPushButton:hover { background: #D0D0D0; font-weight: bold; color: black; }"
                )
            else:
                button.setStyleSheet(
                    "QPushButton { background: #B0B0B0; font-weight: bold; color: black; border: 1px solid #414242; "
                    "border-top-left-radius: 5px; border-top-right-radius: 5px; "
                    "border-bottom-left-radius: 0px; border-bottom-right-radius: 0px; "
                    "border-bottom-color: #414242; font-size: 12pt; }"
                    "QPushButton:hover { background: #D0D0D0; font-weight: bold; color: black; }"
                )

    def setup_stacked_widgets(self):
        # Extrai dados do DataFrame selecionado
        data = self.extract_registro_data()

        # Método para configurar os widgets no StackWidgetManager
        widgets = {
            "Informações Gerais": self.stacked_widget_info(data),
            "Setor Responsável": self.stacked_widget_responsaveis(data),
            "Documentos": self.stacked_widget_documentos(data),
            "Anexos": self.stacked_widget_anexos(data),
        }

        for name, widget in widgets.items():
            self.stack_manager.addWidget(widget)
            widget.setObjectName(name)

    def stacked_widget_responsaveis(self, data):
        # Cria um widget básico para o stack
        frame = QFrame()
        layout = QVBoxLayout()

        # Cria e adiciona o QGroupBox "Dados do Setor Responsável pela Contratação"
        dados_responsavel_contratacao_group = self.create_dados_responsavel_contratacao_group()
        layout.addWidget(dados_responsavel_contratacao_group)

        # Define o layout para o frame
        frame.setLayout(layout)

        return frame

    def stacked_widget_documentos(self, data):
        frame = QFrame()
        layout = QVBoxLayout()

        # Cria e adiciona o QGroupBox "Dados do Setor Responsável pela Contratação"
        botao_documentos = self.create_gerar_documentos_group()
        sigdem_group = self.create_GrupoSIGDEM()
        utilidade_group = self.create_utilidades_group()
        layout.addLayout(botao_documentos)
        layout.addWidget(sigdem_group)
        layout.addLayout(utilidade_group)

        # Define o layout para o frame
        frame.setLayout(layout)        
        return frame
    
    def stacked_widget_anexos(self, data):
        frame = QFrame()
        layout = QVBoxLayout()

        # Cria e adiciona o QGroupBox "Dados do Setor Responsável pela Contratação"
        anexos_group = self.create_anexos_group()
        layout.addWidget(anexos_group)

        # Define o layout para o frame
        frame.setLayout(layout)        
        return frame

    def stacked_widget_info(self, data):
        # Cria um widget básico para o stack
        frame = QFrame()
        layout = QVBoxLayout()

        # Layout horizontal para agrupar os QGroupBox
        hbox_top_layout = QHBoxLayout()

        # Cria e adiciona o QGroupBox "Contratação" ao layout horizontal
        contratacao_group_box = self.create_contratacao_group(data)
        hbox_top_layout.addWidget(contratacao_group_box)

        # Cria um layout vertical para "Classificação Orçamentária" e "Formulário"
        layout_orcamentario_formulario = QVBoxLayout()

        # Cria e adiciona o QGroupBox "Classificação Orçamentária" ao layout vertical
        classificacao_orcamentaria_group_box = self.create_classificacao_orcamentaria_group()
        layout_orcamentario_formulario.addWidget(classificacao_orcamentaria_group_box)

        # Cria o "Formulário de Dados" e adiciona ao layout vertical
        formulario_group_box = self.create_frame_formulario_group()
        layout_orcamentario_formulario.addWidget(formulario_group_box)

        # Adiciona o layout vertical ao layout horizontal
        hbox_top_layout.addLayout(layout_orcamentario_formulario)

        # Adiciona o layout horizontal ao layout principal
        layout.addLayout(hbox_top_layout)
        frame.setLayout(layout)

        return frame

    def create_contratacao_group(self, data):
        contratacao_group_box = QGroupBox("Contratação")
        self.apply_widget_style(contratacao_group_box)
        contratacao_group_box.setFixedWidth(400)
        contratacao_layout = QVBoxLayout()

        # Objeto
        self.objeto_edit = QLineEdit(data['objeto'])

        # Criando um layout horizontal para o campo de entrada de texto e o ícone
        objeto_layout = QHBoxLayout()

        objeto_label = QLabel("Objeto:")
        objeto_layout.addWidget(objeto_label)
        objeto_layout.addWidget(self.objeto_edit)

        # Criando o ícone
        icon_label = QLabel()
        icon = QIcon(str(self.ICONS_DIR / "prioridade.png"))
        icon_pixmap = icon.pixmap(30, 30)  # Definindo o tamanho do ícone
        icon_label.setPixmap(icon_pixmap)

        # Adicionando o ícone ao layout
        objeto_layout.addWidget(icon_label)

        # Aplicando estilo diretamente aos widgets dentro do layout
        self.apply_widget_style(self.objeto_edit)
        self.apply_widget_style(icon_label)

        # Adicionando o layout horizontal diretamente ao layout principal de contratação
        contratacao_layout.addLayout(objeto_layout)

        # Configuração Situação
        situacao_layout = QHBoxLayout()
        situacao_label = QLabel("Situação:")
        self.apply_widget_style(situacao_label)
        self.situacao_edit = self.create_combo_box(data.get('situacao', 'Planejamento'), ["Planejamento", "Aprovado", "Sessão Pública", "Homologado", "Empenhado", "Concluído", "Arquivado"], 185, 35)
        situacao_layout.addWidget(situacao_label)
        situacao_layout.addWidget(self.situacao_edit)
        contratacao_layout.addLayout(situacao_layout)

        # Adiciona outros layouts ao layout de contratação
        self.nup_edit = QLineEdit(data['nup'])
        contratacao_layout.addLayout(self.create_layout("NUP:", self.nup_edit))

        # Configuração de Material/Serviço na mesma linha
        material_layout = QHBoxLayout()
        material_label = QLabel("Material/Serviço:")
        self.apply_widget_style(material_label)
        self.material_edit = self.create_combo_box(data.get('material_servico', 'Material'), ["Material", "Serviço"], 185, 35)
        material_layout.addWidget(material_label)
        material_layout.addWidget(self.material_edit)
        contratacao_layout.addLayout(material_layout)

        # Configuração da Data da Sessão na mesma linha
        data_layout = QHBoxLayout()
        data_label = QLabel("Data da Sessão Pública:")
        self.apply_widget_style(data_label)
        self.data_edit = QDateEdit()
        # self.data_edit.setFixedWidth(120)
        self.data_edit.setCalendarPopup(True)
        data_sessao_str = data.get('data_sessao', '')
        if data_sessao_str:
            self.data_edit.setDate(QDate.fromString(data_sessao_str, "yyyy-MM-dd"))
        else:
            self.data_edit.setDate(QDate.currentDate())
        data_layout.addWidget(data_label)
        data_layout.addWidget(self.data_edit)
        contratacao_layout.addLayout(data_layout)

        previsao_contratacao_layout = QHBoxLayout()
        previsao_contratacao_label = QLabel("Previsão da Contratação:")
        self.apply_widget_style(previsao_contratacao_label)
        self.previsao_contratacao_edit = QDateEdit()
        # self.previsao_contratacao_edit.setFixedWidth(120)
        self.previsao_contratacao_edit.setCalendarPopup(True)
        previsao_contratacao_str = data.get('previsao_contratacao', '')
        if previsao_contratacao_str:
            self.previsao_contratacao_edit.setDate(QDate.fromString(previsao_contratacao_str, "yyyy-MM-dd"))
        else:
            self.previsao_contratacao_edit.setDate(QDate.currentDate())
        previsao_contratacao_layout.addWidget(previsao_contratacao_label)
        previsao_contratacao_layout.addWidget(self.previsao_contratacao_edit)
        contratacao_layout.addLayout(previsao_contratacao_layout)

        # Vigência
        self.vigencia_edit = QComboBox()
        self.vigencia_edit.setEditable(True)
        for i in range(1, 13):
            self.vigencia_edit.addItem(f"{i} ({self.number_to_text(i)}) meses")
        vigencia = data.get('vigencia', '2 (dois) meses')
        self.vigencia_edit.setCurrentText(vigencia)
        contratacao_layout.addLayout(self.create_layout("Vigência:", self.vigencia_edit))

        # Configuração de Critério de Julgamento na mesma linha
        criterio_layout = QHBoxLayout()
        criterio_label = QLabel("Critério Julgamento:")
        self.apply_widget_style(criterio_label)
        self.criterio_edit = self.create_combo_box(data.get('criterio_julgamento', 'Menor Preço'), ["Menor Preço", "Maior Desconto"], 185, 35)
        criterio_layout.addWidget(criterio_label)
        criterio_layout.addWidget(self.criterio_edit)
        contratacao_layout.addLayout(criterio_layout)

        # Configuração de Com Disputa na mesma linha
        disputa_layout = QHBoxLayout()
        disputa_label = QLabel("Com disputa?")
        self.apply_widget_style(disputa_label)
        self.radio_disputa_sim = QRadioButton("Sim")
        self.radio_disputa_nao = QRadioButton("Não")
        self.disputa_group = QButtonGroup(self)
        self.disputa_group.addButton(self.radio_disputa_sim)
        self.disputa_group.addButton(self.radio_disputa_nao)
        disputa_layout.addWidget(disputa_label)
        disputa_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))

        disputa_layout.addWidget(self.radio_disputa_sim)
        disputa_layout.addWidget(self.radio_disputa_nao)
        contratacao_layout.addLayout(disputa_layout)

        com_disputa_value = data.get('com_disputa', 'Sim')
        if com_disputa_value is None or pd.isna(com_disputa_value):
            com_disputa_value = 'Sim'
        self.radio_disputa_sim.setChecked(com_disputa_value == 'Sim')
        self.radio_disputa_nao.setChecked(com_disputa_value != 'Sim')

        # Pesquisa de Preço Concomitante
        pesquisa_concomitante_layout = QHBoxLayout()
        pesquisa_concomitante_label = QLabel("Pesquisa Concomitante?")
        self.apply_widget_style(pesquisa_concomitante_label)
        self.radio_pesquisa_sim = QRadioButton("Sim")
        self.radio_pesquisa_nao = QRadioButton("Não")
        self.pesquisa_group = QButtonGroup(self)
        self.pesquisa_group.addButton(self.radio_pesquisa_sim)
        self.pesquisa_group.addButton(self.radio_pesquisa_nao)
        pesquisa_preco_value = data.get('pesquisa_preco', 'Não')
        self.radio_pesquisa_sim.setChecked(pesquisa_preco_value == 'Sim')
        self.radio_pesquisa_nao.setChecked(pesquisa_preco_value != 'Sim')
        pesquisa_concomitante_layout.addWidget(pesquisa_concomitante_label)
        pesquisa_concomitante_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))

        pesquisa_concomitante_layout.addWidget(self.radio_pesquisa_sim)
        pesquisa_concomitante_layout.addWidget(self.radio_pesquisa_nao)
        contratacao_layout.addLayout(pesquisa_concomitante_layout)

        # Atividade de Custeio
        atividade_custeio_layout = QHBoxLayout()
        custeio_label = QLabel("Atividade de Custeio?")
        self.apply_widget_style(custeio_label)
        self.radio_custeio_sim = QRadioButton("Sim")
        self.radio_custeio_nao = QRadioButton("Não")
        atividade_custeio_value = data.get('atividade_custeio', 'Não')
        self.radio_custeio_sim.setChecked(atividade_custeio_value == 'Sim')
        self.radio_custeio_nao.setChecked(atividade_custeio_value != 'Sim')
        atividade_custeio_layout.addWidget(custeio_label)
        atividade_custeio_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))

        atividade_custeio_layout.addWidget(self.radio_custeio_sim)
        atividade_custeio_layout.addWidget(self.radio_custeio_nao)
        contratacao_layout.addLayout(atividade_custeio_layout)

        # contratacao_layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        # Layout Link PNCP
        link_pncp_layout = QHBoxLayout()

        self.link_pncp_edit = QLineEdit(data.get('link_pncp', ''))
        link_pncp_layout.addLayout(self.create_layout("Link PNCP:", self.link_pncp_edit))

        icon_link = QIcon(str(self.ICONS_DIR / "link.png"))
        link_pncp_button = self.create_button(
            "",
            icon=icon_link,
            callback=self.on_link_pncp_clicked,
            tooltip_text="Clique para acessar o Link da dispensa no Portal Nacional de Contratações Públicas (PNCP)",
            button_size=QSize(30, 30),
            icon_size=QSize(30, 30)
        )
        self.apply_widget_style(link_pncp_button)
        link_pncp_layout.addWidget(link_pncp_button)
        contratacao_layout.addLayout(link_pncp_layout)

        contratacao_group_box.setLayout(contratacao_layout)

        return contratacao_group_box

    def on_link_pncp_clicked(self):
        url = self.link_pncp_edit.text()
        if url:
            webbrowser.open(url)
        else:
            # Exibir uma mensagem de erro caso o campo esteja vazio
            print("URL não encontrada. Por favor, insira um link válido.")

    def create_classificacao_orcamentaria_group(self):
        data = self.extract_registro_data()
        classificacao_orcamentaria_group_box = QGroupBox("Classificação Orçamentária")
        self.apply_widget_style(classificacao_orcamentaria_group_box)
        classificacao_orcamentaria_group_box.setFixedWidth(400)  
        classificacao_orcamentaria_layout = QVBoxLayout()

        # Valor Estimado
        self.valor_edit = RealLineEdit(str(data['valor_total']) if pd.notna(data['valor_total']) else "")
        valor_layout = QHBoxLayout()
        valor_label = QLabel("Valor Estimado:")
        valor_layout.addWidget(valor_label)
        valor_layout.addWidget(self.valor_edit)

        # Criando o ícone
        icon_label_layout = QLabel()
        icon_valor = QIcon(str(self.ICONS_DIR / "emenda_parlamentar.png"))
        icon_pixmap_valor = icon_valor.pixmap(30, 30)  # Definindo o tamanho do ícone
        icon_label_layout.setPixmap(icon_pixmap_valor)

        # Adicionando o ícone ao layout
        valor_layout.addWidget(icon_label_layout)

        classificacao_orcamentaria_layout.addLayout(valor_layout)

        self.acao_interna_edit = QLineEdit(data['acao_interna'])
        self.fonte_recurso_edit = QLineEdit(data['fonte_recursos'])
        self.natureza_despesa_edit = QLineEdit(data['natureza_despesa'])
        self.unidade_orcamentaria_edit = QLineEdit(data['unidade_orcamentaria'])
        self.ptres_edit = QLineEdit(data['programa_trabalho_resuminho'])

        classificacao_orcamentaria_layout.addLayout(self.create_layout("Ação Interna:", self.acao_interna_edit))
        classificacao_orcamentaria_layout.addLayout(self.create_layout("Fonte de Recurso (FR):", self.fonte_recurso_edit))
        classificacao_orcamentaria_layout.addLayout(self.create_layout("Natureza de Despesa (ND):", self.natureza_despesa_edit))
        classificacao_orcamentaria_layout.addLayout(self.create_layout("Unidade Orçamentária (UO):", self.unidade_orcamentaria_edit))
        classificacao_orcamentaria_layout.addLayout(self.create_layout("PTRES:", self.ptres_edit))
        
        # classificacao_orcamentaria_layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        classificacao_orcamentaria_group_box.setLayout(classificacao_orcamentaria_layout)

        return classificacao_orcamentaria_group_box
    
    def update_title_label(self):
        data = self.extract_registro_data()
        html_text = (
            f"{data['tipo']} {data['numero']}/{data['ano']} - {data['objeto']}<br>"
            f"<span style='font-size: 16px'>OM: {data['orgao_responsavel']} (UASG: {data['uasg']})</span>"
        )

        if not hasattr(self, 'titleLabel'):
            self.titleLabel = QLabel()
            self.titleLabel.setTextFormat(Qt.TextFormat.RichText)
            self.titleLabel.setStyleSheet("font-size: 26px; font-weight: bold;")

        self.titleLabel.setText(html_text)

        if not hasattr(self, 'header_layout'):
            self.header_layout = QHBoxLayout()
            self.header_layout.addWidget(self.titleLabel)  # Adiciona o QLabel ao layout

            header_widget = QWidget()
            header_widget.setLayout(self.header_layout)
            header_widget.setFixedHeight(80)
            self.header_widget = header_widget

        return self.header_widget

    def update_title_label_text(self, new_title):
        data = self.extract_registro_data()
        html_text = (
            f"{data['tipo']} {data['numero']}/{data['ano']} - {data['objeto']}<br>"
            f"<span style='font-size: 16px'>OM: {new_title}</span>"
        )
        self.titleLabel.setText(html_text)
        print(f"Title updated: {html_text}")
    
    def add_action_buttons(self, layout):
        icon_confirm = QIcon(str(self.ICONS_DIR / "confirm.png"))
        
        button_confirm = self.create_button(" Salvar", icon_confirm, self.save_changes, "Salvar dados", QSize(110, 30), QSize(30, 30))
                
        layout.addWidget(button_confirm)

        self.apply_widget_style(button_confirm)
    
    def create_group_box(self, title, details):
        group_box = QGroupBox(title)
        layout = QVBoxLayout()
        layout.addWidget(QLabel(details))
        group_box.setLayout(layout)
        return group_box

    def stacked_widget(self):
        # Preenche o formulário de documentos
        frame = QFrame()
        layout = QVBoxLayout()
        layout.addWidget(QLabel())
        frame.setLayout(layout)
        return frame
    
    def create_button(self, text="", icon=None, callback=None, tooltip_text="", button_size=None, icon_size=None):
        btn = QPushButton(text)
        if icon:
            btn.setIcon(icon)
            btn.setIconSize(icon_size if icon_size else QSize(40, 40))
        if button_size:
            btn.setFixedSize(button_size)
        btn.setToolTip(tooltip_text)
        if callback:
            btn.clicked.connect(callback)  # Conecta o callback ao evento de clique
        return btn
    
    def extract_registro_data(self):
        # Verifica se o DataFrame não está vazio
        if self.df_registro_selecionado.empty:
            print("DataFrame está vazio")
            return {}

        # Extrai dados do registro selecionado e armazena como atributos de instância
        self.id_processo = self.df_registro_selecionado['id_processo'].iloc[0]
        self.tipo = self.df_registro_selecionado['tipo'].iloc[0]
        self.numero = self.df_registro_selecionado['numero'].iloc[0]
        self.ano = self.df_registro_selecionado['ano'].iloc[0]
        self.situacao = self.df_registro_selecionado['situacao'].iloc[0]
        self.nup = self.df_registro_selecionado['nup'].iloc[0]
        self.material_servico = self.df_registro_selecionado['material_servico'].iloc[0]
        self.objeto = self.df_registro_selecionado['objeto'].iloc[0]
        self.vigencia = self.df_registro_selecionado['vigencia'].iloc[0]
        self.data_sessao = self.df_registro_selecionado['data_sessao'].iloc[0]
        self.operador = self.df_registro_selecionado['operador'].iloc[0]
        self.criterio_julgamento = self.df_registro_selecionado['criterio_julgamento'].iloc[0]
        self.com_disputa = self.df_registro_selecionado['com_disputa'].iloc[0]
        self.pesquisa_preco = self.df_registro_selecionado['pesquisa_preco'].iloc[0]
        self.previsao_contratacao = self.df_registro_selecionado['previsao_contratacao'].iloc[0]
        self.uasg = self.df_registro_selecionado['uasg'].iloc[0]
        self.orgao_responsavel = self.df_registro_selecionado['orgao_responsavel'].iloc[0]
        self.sigla_om = self.df_registro_selecionado['sigla_om'].iloc[0]
        self.setor_responsavel = self.df_registro_selecionado['setor_responsavel'].iloc[0]
        self.responsavel_pela_demanda = self.df_registro_selecionado['responsavel_pela_demanda'].iloc[0]
        self.ordenador_despesas = self.df_registro_selecionado['ordenador_despesas'].iloc[0]
        self.agente_fiscal = self.df_registro_selecionado['agente_fiscal'].iloc[0]
        self.gerente_de_credito = self.df_registro_selecionado['gerente_de_credito'].iloc[0]
        self.cod_par = self.df_registro_selecionado['cod_par'].iloc[0]
        self.prioridade_par = self.df_registro_selecionado['prioridade_par'].iloc[0]
        self.cep = self.df_registro_selecionado['cep'].iloc[0]
        self.endereco = self.df_registro_selecionado['endereco'].iloc[0]
        self.email = self.df_registro_selecionado['email'].iloc[0]
        self.telefone = self.df_registro_selecionado['telefone'].iloc[0]
        self.dias_para_recebimento = self.df_registro_selecionado['dias_para_recebimento'].iloc[0]
        self.horario_para_recebimento = self.df_registro_selecionado['horario_para_recebimento'].iloc[0]
        self.valor_total = self.df_registro_selecionado['valor_total'].iloc[0]
        self.acao_interna = self.df_registro_selecionado['acao_interna'].iloc[0]
        self.fonte_recursos = self.df_registro_selecionado['fonte_recursos'].iloc[0]
        self.natureza_despesa = self.df_registro_selecionado['natureza_despesa'].iloc[0]
        self.unidade_orcamentaria = self.df_registro_selecionado['unidade_orcamentaria'].iloc[0]
        self.programa_trabalho_resuminho = self.df_registro_selecionado['programa_trabalho_resuminho'].iloc[0]
        self.atividade_custeio = self.df_registro_selecionado['atividade_custeio'].iloc[0]
        self.comentarios = self.df_registro_selecionado['comentarios'].iloc[0]
        self.justificativa = self.df_registro_selecionado['justificativa'].iloc[0]
        self.link_pncp = self.df_registro_selecionado['link_pncp'].iloc[0]
        self.comunicacao_padronizada = self.df_registro_selecionado['comunicacao_padronizada'].iloc[0]

        data = {
            'id_processo': self.id_processo,
            'tipo': self.tipo,
            'numero': self.numero,
            'ano': self.ano,
            'situacao': self.situacao,
            'nup': self.nup,
            'material_servico': self.material_servico,
            'objeto': self.objeto,
            'vigencia': self.vigencia,
            'data_sessao': self.data_sessao,
            'operador': self.operador,
            'criterio_julgamento': self.criterio_julgamento,
            'com_disputa': self.com_disputa,
            'pesquisa_preco': self.pesquisa_preco,
            'previsao_contratacao': self.previsao_contratacao,
            'uasg': self.uasg,
            'orgao_responsavel': self.orgao_responsavel,
            'sigla_om': self.sigla_om,
            'setor_responsavel': self.setor_responsavel,
            'responsavel_pela_demanda': self.responsavel_pela_demanda,
            'ordenador_despesas': self.ordenador_despesas,
            'agente_fiscal': self.agente_fiscal,
            'gerente_de_credito': self.gerente_de_credito,
            'cod_par': self.cod_par,
            'prioridade_par': self.prioridade_par,
            'cep': self.cep,
            'endereco': self.endereco,
            'email': self.email,
            'telefone': self.telefone,
            'dias_para_recebimento': self.dias_para_recebimento,
            'horario_para_recebimento': self.horario_para_recebimento,
            'valor_total': self.valor_total,
            'acao_interna': self.acao_interna,
            'fonte_recursos': self.fonte_recursos,
            'natureza_despesa': self.natureza_despesa,
            'unidade_orcamentaria': self.unidade_orcamentaria,
            'programa_trabalho_resuminho': self.programa_trabalho_resuminho,
            'atividade_custeio': self.atividade_custeio,
            'comentarios': self.comentarios,
            'justificativa': self.justificativa,
            'link_pncp': self.link_pncp,
            'comunicacao_padronizada': self.comunicacao_padronizada,
        }

        return data

    def save_changes(self):
        try:
            data = {
                'situacao': self.situacao_edit.currentText(),
                'ordenador_despesas': self.ordenador_combo.currentText(),
                'agente_fiscal': self.agente_fiscal_combo.currentText(),
                'gerente_de_credito': self.gerente_credito_combo.currentText(),
                'responsavel_pela_demanda': self.responsavel_demanda_combo.currentText(),
                'nup': self.nup_edit.text().strip(),
                'material_servico': self.material_edit.currentText(),
                'objeto': self.objeto_edit.text().strip(),
                'vigencia': self.vigencia_edit.currentText(),
                'data_sessao': self.data_edit.date().toString("yyyy-MM-dd"),
                'previsao_contratacao': self.previsao_contratacao_edit.date().toString("yyyy-MM-dd"),
                'criterio_julgamento': self.criterio_edit.currentText(),
                'com_disputa': 'Sim' if self.radio_disputa_sim.isChecked() else 'Não',
                'pesquisa_preco': 'Sim' if self.radio_pesquisa_sim.isChecked() else 'Não',
                'setor_responsavel': self.setor_responsavel_combo.currentText(),
                'operador': self.operador_dispensa_combo.currentText(),
                'sigla_om': self.om_combo.currentText(),
                'uasg': self.df_registro_selecionado.at[self.df_registro_selecionado.index[0], 'uasg'],
                'orgao_responsavel': self.df_registro_selecionado.at[self.df_registro_selecionado.index[0], 'orgao_responsavel'],
                'cod_par': self.par_edit.text().strip(),
                'prioridade_par': self.prioridade_combo.currentText(),
                'cep': self.cep_edit.text().strip(),
                'endereco': self.endereco_edit.text().strip(),
                'email': self.email_edit.text().strip(),
                'telefone': self.telefone_edit.text().strip(),
                'dias_para_recebimento': self.dias_edit.text().strip(),
                'horario_para_recebimento': self.horario_edit.text().strip(),
                'justificativa': self.justificativa_edit.toPlainText().strip(),
                'valor_total': self.valor_edit.text().strip(),
                'acao_interna': self.acao_interna_edit.text().strip(),
                'fonte_recursos': self.fonte_recurso_edit.text().strip(),
                'natureza_despesa': self.natureza_despesa_edit.text().strip(),
                'unidade_orcamentaria': self.unidade_orcamentaria_edit.text().strip(),
                'programa_trabalho_resuminho': self.ptres_edit.text().strip(),
                'atividade_custeio': 'Sim' if self.radio_custeio_sim.isChecked() else 'Não',
                'comunicacao_padronizada': self.cp_edit.text().strip(),
            }

            # Atualizar o DataFrame com os novos valores
            for key, value in data.items():
                self.df_registro_selecionado.at[self.df_registro_selecionado.index[0], key] = value

            # Atualizar banco de dados
            self.update_database(data)
            self.dados_atualizados.emit()

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao salvar as alterações: {str(e)}")

    def update_database(self, data):
        with self.database_manager as connection:
            cursor = connection.cursor()
            set_part = ', '.join([f"{key} = ?" for key in data.keys()])
            valores = list(data.values())
            valores.append(self.df_registro_selecionado['id_processo'].iloc[0])
            query = f"UPDATE controle_dispensas SET {set_part} WHERE id_processo = ?"
            cursor.execute(query, valores)
            connection.commit()
            QMessageBox.information(self, "Atualização", "As alterações foram salvas com sucesso.")

    def fill_frame_dados_do_setor_resposavel_contratacao(self):
        frame = QFrame(self)
        layout = QVBoxLayout()

        label_setor = QLabel("Setor Responsável")
        layout.addWidget(label_setor)

        frame.setLayout(layout)
        return frame

    def apply_widget_style(self, widget):
        widget.setStyleSheet("font-size: 11pt;") 

    def apply_widget_style_11(self, widget):
        widget.setStyleSheet("font-size: 11pt;") 

    def apply_widget_style_12(self, widget):
        widget.setStyleSheet("font-size: 12pt;") 

    def apply_widget_style_14(self, widget):
        widget.setStyleSheet("font-size: 14pt;") 

    def create_combo_box(self, current_text, items, fixed_width, fixed_height):
        combo_box = QComboBox()
        combo_box.addItems(items)
        combo_box.setFixedWidth(fixed_width)
        combo_box.setFixedHeight(fixed_height)  # Define a altura fixa do ComboBox
        combo_box.setStyleSheet("QComboBox { font-size: 12px; }")  # Ajusta o estilo para melhor visualização
        self.apply_widget_style(combo_box)
        combo_box.setCurrentText(current_text)
        return combo_box

    def create_layout(self, label_text, widget, fixed_width=None):
        layout = QHBoxLayout()
        label = QLabel(label_text)
        self.apply_widget_style(label)
        
        # Adiciona a largura fixa se especificada
        if fixed_width and isinstance(widget, QWidget):
            widget.setFixedWidth(fixed_width)
        
        # Aplica estilo apenas se o widget for uma instância de QWidget
        if isinstance(widget, QWidget):
            self.apply_widget_style(widget)
        
        layout.addWidget(label)
        layout.addWidget(widget)
        return layout
    
    def number_to_text(self, number):
        numbers_in_words = ["um", "dois", "três", "quatro", "cinco", "seis", "sete", "oito", "nove", "dez", "onze", "doze"]
        return numbers_in_words[number - 1] 

    def create_frame_formulario_group(self):
        formulario_group_box = QGroupBox("Formulário de Dados")
        self.apply_widget_style(formulario_group_box)   
        formulario_group_box.setFixedWidth(400)   
        formulario_layout = QVBoxLayout()


        # Adicionando os botões ao layout
        icon_excel_up = QIcon(str(self.ICONS_DIR / "excel_up.png"))
        icon_excel_down = QIcon(str(self.ICONS_DIR / "excel_down.png"))
        icon_standard = QIcon(str(self.ICONS_DIR / "standard.png"))

        criar_formulario_button = self.create_button(
            "   Criar Formulário   ", 
            icon=icon_excel_up, 
            callback=self.formulario_excel.criar_formulario, 
            tooltip_text="Clique para criar o formulário", 
            button_size=QSize(220, 50), 
            icon_size=QSize(45, 45)
        )

        carregar_formulario_button = self.create_button(
            "Carregar Formulário", 
            icon=icon_excel_down, 
            callback=self.formulario_excel.carregar_formulario, 
            tooltip_text="Clique para carregar o formulário", 
            button_size=QSize(220, 50), 
            icon_size=QSize(45, 45)
        )

        visualizar_pdf_button = self.create_button(
            "      Pré-Definições     ",
            icon=icon_standard,
            callback=self.selecionar_predefinicoes, 
            tooltip_text="Clique para alterar ou escolher os dados predefinidos", 
            button_size=QSize(220, 40), 
            icon_size=QSize(30, 30)
        )       

        formulario_layout.addWidget(criar_formulario_button, alignment=Qt.AlignmentFlag.AlignCenter)
        formulario_layout.addWidget(carregar_formulario_button, alignment=Qt.AlignmentFlag.AlignCenter)
        # formulario_layout.addWidget(visualizar_pdf_button, alignment=Qt.AlignmentFlag.AlignCenter)
        formulario_group_box.setLayout(formulario_layout)

        return formulario_group_box
    
    def selecionar_predefinicoes(self):
        pass

    def preencher_campos(self):
        try:
            self.situacao_edit.setCurrentText(str(self.df_registro_selecionado.at[0, 'situacao']))
            self.ordenador_combo.setCurrentText(str(self.df_registro_selecionado.at[0, 'ordenador_despesas']))
            self.agente_fiscal_combo.setCurrentText(str(self.df_registro_selecionado.at[0, 'agente_fiscal']))
            self.gerente_credito_combo.setCurrentText(str(self.df_registro_selecionado.at[0, 'gerente_de_credito']))
            self.responsavel_demanda_combo.setCurrentText(str(self.df_registro_selecionado.at[0, 'responsavel_pela_demanda']))
            self.nup_edit.setText(str(self.df_registro_selecionado.at[0, 'nup']))
            self.material_edit.setCurrentText(str(self.df_registro_selecionado.at[0, 'material_servico']))
            self.objeto_edit.setText(str(self.df_registro_selecionado.at[0, 'objeto']))
            self.vigencia_edit.setCurrentText(str(self.df_registro_selecionado.at[0, 'vigencia']))
            self.data_edit.setDate(QDate.fromString(str(self.df_registro_selecionado.at[0, 'data_sessao']), "yyyy-MM-dd"))
            self.previsao_contratacao_edit.setDate(QDate.fromString(str(self.df_registro_selecionado.at[0, 'previsao_contratacao']), "yyyy-MM-dd"))
            self.criterio_edit.setCurrentText(str(self.df_registro_selecionado.at[0, 'criterio_julgamento']))
            self.radio_disputa_sim.setChecked(str(self.df_registro_selecionado.at[0, 'com_disputa']) == 'Sim')
            self.radio_disputa_nao.setChecked(str(self.df_registro_selecionado.at[0, 'com_disputa']) == 'Não')
            self.radio_pesquisa_sim.setChecked(str(self.df_registro_selecionado.at[0, 'pesquisa_preco']) == 'Sim')
            self.radio_pesquisa_nao.setChecked(str(self.df_registro_selecionado.at[0, 'pesquisa_preco']) == 'Não')
            self.setor_responsavel_combo.setCurrentText(str(self.df_registro_selecionado.at[0, 'setor_responsavel']))
            self.operador_dispensa_combo.setCurrentText(str(self.df_registro_selecionado.at[0, 'operador']))
            self.om_combo.setCurrentText(str(self.df_registro_selecionado.at[0, 'sigla_om']))
            self.par_edit.setText(str(self.df_registro_selecionado.at[0, 'cod_par']))
            self.prioridade_combo.setCurrentText(str(self.df_registro_selecionado.at[0, 'prioridade_par']))
            self.cep_edit.setText(str(self.df_registro_selecionado.at[0, 'cep']))
            self.endereco_edit.setText(str(self.df_registro_selecionado.at[0, 'endereco']))
            self.email_edit.setText(str(self.df_registro_selecionado.at[0, 'email']))
            self.telefone_edit.setText(str(self.df_registro_selecionado.at[0, 'telefone']))
            self.dias_edit.setText(str(self.df_registro_selecionado.at[0, 'dias_para_recebimento']))
            self.horario_edit.setText(str(self.df_registro_selecionado.at[0, 'horario_para_recebimento']))
            self.justificativa_edit.setPlainText(str(self.df_registro_selecionado.at[0, 'justificativa']))
            self.valor_edit.setText(str(self.df_registro_selecionado.at[0, 'valor_total']))
            self.acao_interna_edit.setText(str(self.df_registro_selecionado.at[0, 'acao_interna']))
            self.fonte_recurso_edit.setText(str(self.df_registro_selecionado.at[0, 'fonte_recursos']))
            self.natureza_despesa_edit.setText(str(self.df_registro_selecionado.at[0, 'natureza_despesa']))
            self.unidade_orcamentaria_edit.setText(str(self.df_registro_selecionado.at[0, 'unidade_orcamentaria']))
            self.ptres_edit.setText(str(self.df_registro_selecionado.at[0, 'programa_trabalho_resuminho']))
            self.radio_custeio_sim.setChecked(str(self.df_registro_selecionado.at[0, 'atividade_custeio']) == 'Sim')
            self.radio_custeio_nao.setChecked(str(self.df_registro_selecionado.at[0, 'atividade_custeio']) == 'Não')
            self.cp_edit.setText(str(self.df_registro_selecionado.at[0, 'comunicacao_padronizada']))

        except KeyError as e:
            print(f"Erro ao preencher campos: {str(e)}")
    
    """
    
    
    
    



    
    PARTE REFERENTE AS UTILIDADES
    
    






    """

    def create_utilidades_group(self):
        utilidades_layout = QHBoxLayout()
        utilidades_layout.setSpacing(0)
        utilidades_layout.setContentsMargins(0, 0, 0, 0)

        # Botão para abrir o arquivo de registro
        icon_salvar_pasta = QIcon(str(self.ICONS_DIR / "salvar_pasta.png"))
        editar_registro_button = self.create_button("Local de Salvamento", icon=icon_salvar_pasta, callback=self.consolidador.alterar_diretorio_base, tooltip_text="Clique para alterar o local de salvamento dos arquivos", button_size=QSize(210, 40), icon_size=QSize(40, 40))
        self.apply_widget_style(editar_registro_button)
        utilidades_layout.addWidget(editar_registro_button, alignment=Qt.AlignmentFlag.AlignCenter)

        # Botão para abrir o arquivo de registro
        icon_open_folder = QIcon(str(self.ICONS_DIR / "open-folder.png"))
        visualizar_pdf_button = self.create_button("Abrir Pasta Base", icon=icon_open_folder, callback=self.consolidador.abrir_pasta_base, tooltip_text="Clique para alterar ou escolher os dados predefinidos", button_size=QSize(210, 40), icon_size=QSize(40, 40))
        self.apply_widget_style(visualizar_pdf_button)
        utilidades_layout.addWidget(visualizar_pdf_button, alignment=Qt.AlignmentFlag.AlignCenter)

        # Botão para abrir o arquivo de registro
        icon_template = QIcon(str(self.ICONS_DIR / "template.png"))
        visualizar_pdf_button = self.create_button("Editar Modelos", icon=icon_template, callback=self.consolidador.editar_modelo, tooltip_text="Clique para editar os modelos dos documentos", button_size=QSize(210, 40), icon_size=QSize(40, 40))
        self.apply_widget_style(visualizar_pdf_button)
        utilidades_layout.addWidget(visualizar_pdf_button, alignment=Qt.AlignmentFlag.AlignCenter)

        icon_info_sigdem = QIcon(str(self.ICONS_DIR / "info_sigdem.png"))
        info_sigdem_button = self.create_button("Informações SIGDEM", icon=icon_info_sigdem, callback=self.on_autorizacao_clicked, tooltip_text="Clique para gerar a Declaração de Adequação Orçamentária", button_size=QSize(210, 40), icon_size=QSize(40, 40))
        self.apply_widget_style(info_sigdem_button)
        utilidades_layout.addWidget(info_sigdem_button, alignment=Qt.AlignmentFlag.AlignCenter)

        return utilidades_layout
    
    """
    
    
    
    



    
    PARTE REFERENTE AO GROUPO DOS AGENTES RESPONSAVEIS
    
    






    """

    def create_agentes_responsaveis_layout(self):
        # Frame para agentes responsáveis com borda lateral
        frame_agentes = QFrame()
        # Criação do layout principal para os agentes responsáveis
        agente_responsavel_layout = QVBoxLayout(frame_agentes)
        agente_responsavel_layout.setContentsMargins(10, 1, 10, 1)  # Define margens ao redor do layout

        # Criação dos ComboBox com ajuste de altura
        self.ordenador_combo = self.create_combo_box('', [], 260, 70)
        self.agente_fiscal_combo = self.create_combo_box('', [], 260, 70)
        self.gerente_credito_combo = self.create_combo_box('', [], 260, 65)
        self.responsavel_demanda_combo = self.create_combo_box('', [], 260, 65)
        self.operador_dispensa_combo = self.create_combo_box('', [], 260, 70)

        # Adicionando labels e ComboBox diretamente ao layout
        labels_combos = [
            ("Ordenador de Despesa:", self.ordenador_combo),
            ("Agente Fiscal:", self.agente_fiscal_combo),
            ("Gerente de Crédito:", self.gerente_credito_combo),
            ("Responsável pela Demanda:", self.responsavel_demanda_combo),
            ("Operador da Contratação:", self.operador_dispensa_combo)
        ]

        for label_text, combo_box in labels_combos:
            # Cria um layout horizontal para a label e o ComboBox
            h_layout = QVBoxLayout()
            h_layout.setSpacing(0)  # Ajusta o espaçamento entre label e ComboBox
            h_layout.setContentsMargins(0, 0, 0, 0)  # Margens para o layout

            # Cria e estiliza a label
            label = QLabel(label_text)
            label.setStyleSheet("color: #8AB4F7; font-size: 16px")
            label.setContentsMargins(0, 0, 0, 0)  # Define margens para a label

            # Adiciona a label e o ComboBox ao layout horizontal
            h_layout.addWidget(label)
            h_layout.addWidget(combo_box)

            # Adiciona o layout horizontal ao layout principal
            agente_responsavel_layout.addLayout(h_layout)

        # Carrega os agentes responsáveis para popular os ComboBoxes
        self.carregarAgentesResponsaveis()

        return frame_agentes


    def carregarAgentesResponsaveis(self):
        try:
            print("Tentando conectar ao banco de dados...")
            with sqlite3.connect(self.database_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='controle_agentes_responsaveis'")
                if cursor.fetchone() is None:
                    raise Exception("A tabela 'controle_agentes_responsaveis' não existe no banco de dados. Configure os Ordenadores de Despesa no Módulo 'Configurações'.")

                print("Tabela 'controle_agentes_responsaveis' encontrada. Carregando dados...")
                self.carregarDadosCombo(conn, cursor, "Ordenador de Despesa%", self.ordenador_combo)
                self.carregarDadosCombo(conn, cursor, "Agente Fiscal%", self.agente_fiscal_combo)
                self.carregarDadosCombo(conn, cursor, "Gerente de Crédito%", self.gerente_credito_combo)
                self.carregarDadosCombo(conn, cursor, "Operador%", self.operador_dispensa_combo)
                self.carregarDadosCombo(conn, cursor, "NOT LIKE", self.responsavel_demanda_combo)

                # print("Valores carregados no ComboBox:", self.ordenador_combo.count(), "itens")
                # print("Valores carregados no ComboBox:", self.agente_fiscal_combo.count(), "itens")
                # print("Valores carregados no ComboBox:", self.gerente_credito_combo.count(), "itens")
                # print("Valores carregados no ComboBox:", self.operador_dispensa_combo.count(), "itens")
                # print("Valores carregados no ComboBox:", self.responsavel_demanda_combo.count(), "itens")

                # Preencher comboboxes com os valores de df_registro_selecionado se disponíveis
                self.preencher_campos()

        except Exception as e:
            print(f"Erro ao carregar Ordenadores de Despesas: {e}")

    def preencher_combobox_selecionado(self, combo_widget, coluna):
        valor = self.df_registro_selecionado.get(coluna)
        if valor:
            index = combo_widget.findText(valor)
            if index != -1:
                combo_widget.setCurrentIndex(index)
                
    def carregarDadosCombo(self, conn, cursor, funcao_like, combo_widget):
        if "NOT LIKE" in funcao_like:
            sql_query = """
                SELECT nome, posto, funcao FROM controle_agentes_responsaveis
                WHERE funcao NOT LIKE 'Ordenador de Despesa%' AND
                    funcao NOT LIKE 'Agente Fiscal%' AND
                    funcao NOT LIKE 'Gerente de Crédito%' AND
                    funcao NOT LIKE 'Operador%'
            """
        else:
            sql_query = f"SELECT nome, posto, funcao FROM controle_agentes_responsaveis WHERE funcao LIKE '{funcao_like}'"
        
        agentes_df = pd.read_sql_query(sql_query, conn)
        combo_widget.clear()
        for index, row in agentes_df.iterrows():
            texto_display = f"{row['nome']}\n{row['posto']}\n{row['funcao']}"
            combo_widget.addItem(texto_display, userData=row.to_dict())
            print(f"Valores carregados no ComboBox: {combo_widget.count()} itens")

    def create_anexos_group(self):
        data = self.extract_registro_data()

        # LineEdit para o ID de Dispensa Eletrônica
        self.id_dispensa_eletronica = data.get('id_processo', '')
        id_display = self.id_dispensa_eletronica if self.id_dispensa_eletronica else 'ID não disponível'

        # GroupBox para Anexos
        anexos_group_box = QGroupBox(f"Anexos da {id_display}")
        self.apply_widget_style(anexos_group_box)

        # Layout principal do GroupBox
        anexo_layout = QVBoxLayout()
        
        self.anexos_dict = {}

        # Função auxiliar para adicionar seções de anexos
        def add_anexo_section(section_title, *anexos):
            section_label = QLabel(section_title)
            self.apply_widget_style_14(section_label)
            anexo_layout.addWidget(section_label)
            self.anexos_dict[section_title] = []

            for anexo in anexos:
                layout = QHBoxLayout()

                # Caminho e tooltip
                pasta_anexo = self.define_pasta_anexo(section_title, anexo)
                tooltip_text = self.define_tooltip_text(section_title, anexo)

                # Verificação de arquivo PDF
                icon_label = QLabel()
                icon = self.get_icon_for_anexo(pasta_anexo)
                icon_label.setPixmap(icon.pixmap(QSize(30, 30)))
                layout.addWidget(icon_label)
                layout.addSpacerItem(QSpacerItem(10, 0, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum))

                # Botão para abrir a pasta
                btnabrirpasta = self.create_open_folder_button(pasta_anexo, tooltip_text)
                layout.addWidget(btnabrirpasta)

                # Label do anexo
                anexo_label = QLabel(anexo)
                self.apply_widget_style_12(anexo_label)
                layout.addWidget(anexo_label)
                layout.addSpacerItem(QSpacerItem(10, 0, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum))
                layout.addStretch()

                self.anexos_dict[section_title].append((anexo, icon_label))
                anexo_layout.addLayout(layout)

        # Adiciona seções de anexos
        add_anexo_section("Documento de Formalização de Demanda (DFD)", "Anexo A - Relatório do Safin", "Anexo B - Especificações")
        add_anexo_section("Termo de Referência (TR)", "Anexo - Pesquisa de Preços")
        add_anexo_section("Declaração de Adequação Orçamentária", "Anexo - Relatório do PDM/CATSER")

        justificativa_label = QLabel("Justificativas relevantes")
        justificativa_label.setStyleSheet("font-size: 14pt;")  # Ajuste do tamanho da fonte
        anexo_layout.addWidget(justificativa_label)

        # Botões de Ação
        self.add_buttons_to_layout(anexo_layout)

        # Definição do layout final e do GroupBox
        anexos_group_box.setLayout(anexo_layout)

        return anexos_group_box

    def define_pasta_anexo(self, section_title, anexo):
        """Define o caminho da pasta de anexo baseado no título da seção e nome do anexo."""
        id_processo_modificado = self.id_processo.replace("/", "-")
        objeto_modificado = self.objeto.replace("/", "-")

        if section_title == "Documento de Formalização de Demanda (DFD)":
            if "Anexo A" in anexo:
                return self.pasta_base / f'{id_processo_modificado} - {objeto_modificado}' / '2. CP e anexos' / 'DFD' / 'Anexo A - Relatorio Safin'
            elif "Anexo B" in anexo:
                return self.pasta_base / f'{id_processo_modificado} - {objeto_modificado}' / '2. CP e anexos' / 'DFD' / 'Anexo B - Especificações e Quantidade'
        elif section_title == "Termo de Referência (TR)":
            return self.pasta_base / f'{id_processo_modificado} - {objeto_modificado}' / '2. CP e anexos' / 'TR' / 'Pesquisa de Preços'
        elif section_title == "Declaração de Adequação Orçamentária":
            return self.pasta_base / f'{id_processo_modificado} - {objeto_modificado}' / '2. CP e anexos' / 'Declaracao de Adequação Orçamentária' / 'Relatório do PDM-Catser'
        return None

    def define_tooltip_text(self, section_title, anexo):
        """Retorna o texto da tooltip para um anexo."""
        if section_title == "Documento de Formalização de Demanda (DFD)":
            if "Anexo A" in anexo:
                return "Abrir pasta Anexo A - Relatório do Safin"
            elif "Anexo B" in anexo:
                return "Abrir pasta Anexo B - Especificações e Quantidade"
        elif section_title == "Termo de Referência (TR)":
            return "Abrir pasta Pesquisa de Preços"
        elif section_title == "Declaração de Adequação Orçamentária":
            return "Abrir pasta Relatório do PDM-Catser"
        return "Abrir pasta"

    def get_icon_for_anexo(self, pasta_anexo):
        """Retorna o ícone correto baseado na existência de arquivos PDF."""
        icon_confirm = QIcon(str(self.ICONS_DIR / "confirm_green.png"))
        icon_x = QIcon(str(self.ICONS_DIR / "cancel.png"))
        if pasta_anexo and self.verificar_arquivo_pdf(pasta_anexo):
            return icon_confirm
        return icon_x

    def create_open_folder_button(self, pasta_anexo, tooltip_text):
        """Cria um botão para abrir a pasta com o tooltip especificado."""
        icon_abrir_pasta = QIcon(str(self.ICONS_DIR / "open-folder.png"))
        btnabrirpasta = self.create_button(
            "", icon=icon_abrir_pasta, callback=lambda _, p=pasta_anexo: self.abrir_pasta(p),
            tooltip_text=tooltip_text, button_size=QSize(25, 25), icon_size=QSize(25, 25)
        )
        btnabrirpasta.setToolTipDuration(0)
        return btnabrirpasta

    def add_buttons_to_layout(self, layout):
        """Adiciona os botões de 'Visualizar Anexos' e 'Atualizar Pastas' ao layout."""
        icon_browser = QIcon(str(self.ICONS_DIR / "browser.png"))
        add_pdf_button = self.create_button(
            " Visualizar Anexos",
            icon_browser,
            self.add_pdf_to_merger,
            "Visualizar anexos PDFs",
            QSize(220, 40), QSize(30, 30)
        )

        atualizar_button = self.create_button(
            "   Atualizar Pastas  ",
            QIcon(str(self.ICONS_DIR / "refresh.png")),
            self.atualizar_action,
            "Atualizar os dados",
            QSize(220, 40), QSize(30, 30)
        )

        button_layout_anexo = QHBoxLayout()
        button_layout_anexo.addStretch()
        button_layout_anexo.addWidget(add_pdf_button)
        button_layout_anexo.addStretch()

        button_layout_atualizar = QHBoxLayout()
        button_layout_atualizar.addStretch()
        button_layout_atualizar.addWidget(atualizar_button)
        button_layout_atualizar.addStretch()

        layout.addLayout(button_layout_anexo)
        layout.addLayout(button_layout_atualizar)

    def create_gerar_documentos_group(self):
        gerar_documentos_layout = QVBoxLayout()

        icon_pdf = QIcon(str(self.ICONS_DIR / "pdf.png"))
        icon_copy = QIcon(str(self.ICONS_DIR / "copy.png"))

        buttons_info = [
            ("          Autorização para Abertura      ", self.handle_gerar_autorizacao, self.handle_gerar_autorizacao_sidgem),
            (" Comunicação Padronizada e anexos", self.handle_gerar_comunicacao_padronizada, self.handle_gerar_comunicacao_padronizada_sidgem),
            ("              Aviso de Dispensa               ", self.handle_gerar_aviso_dispensa, self.handle_gerar_aviso_dispensa_sidgem)
        ]

        for text, visualizar_callback, sigdem_callback in buttons_info:
            button_layout = QHBoxLayout()

            visualizar_pdf_button = self.create_button(
                text,
                icon=icon_pdf,
                callback=visualizar_callback,
                tooltip_text="Clique para visualizar o PDF",
                button_size=QSize(310, 40),
                icon_size=QSize(40, 40)
            )
            self.apply_widget_style(visualizar_pdf_button)

            sigdem_button = self.create_button(
                "",
                icon=icon_copy,
                callback=sigdem_callback,
                tooltip_text="Clique para copiar",
                button_size=QSize(40, 40),
                icon_size=QSize(30, 30)
            )
            self.apply_widget_style(sigdem_button)

            button_layout.addWidget(visualizar_pdf_button)
            button_layout.addWidget(sigdem_button)

            gerar_documentos_layout.addLayout(button_layout)

        return gerar_documentos_layout
    
    def handle_gerar_autorizacao(self):
        self.assunto_text = f"{self.id_processo} - Abertura de Dispensa Eletrônica"
        self.sinopse_text = (
            f"Termo de Abertura referente à {self.tipo} nº {self.numero}/{self.ano}, para {self.get_descricao_servico()} {self.objeto}\n"
            f"Processo Administrativo NUP: {self.nup}\n"
            f"Setor Demandante: {self.setor_responsavel}"
        )
        self.update_text_fields()
        self.consolidador.gerar_autorizacao()

    def handle_gerar_autorizacao_sidgem(self):
        self.assunto_text = f"{self.id_processo} - Abertura de Dispensa Eletrônica"
        self.sinopse_text = (
            f"Termo de Abertura referente à {self.tipo} nº {self.numero}/{self.ano}, para {self.get_descricao_servico()} {self.objeto}\n"
            f"Processo Administrativo NUP: {self.nup}\n"
            f"Setor Demandante: {self.setor_responsavel}"
        )
        self.update_text_fields()

    def handle_gerar_comunicacao_padronizada(self):
        self.assunto_text = f"{self.id_processo} - CP e Anexos"
        self.sinopse_text = (
            f"Documentos de Planejamento (DFD, TR e Declaração de Adequação Orçamentária) referente à {self.tipo} nº {self.numero}/{self.ano}, para {self.get_descricao_servico()} {self.objeto}\n"
            f"Processo Administrativo NUP: {self.nup}\n"
            f"Setor Demandante: {self.setor_responsavel}"
        )
        self.update_text_fields()
        self.consolidador.gerar_comunicacao_padronizada()

    def handle_gerar_comunicacao_padronizada_sidgem(self):
        self.assunto_text = f"{self.id_processo} - CP e Anexos"
        self.sinopse_text = (
            f"Documentos de Planejamento (DFD, TR e Declaração de Adequação Orçamentária) referente à {self.tipo} nº {self.numero}/{self.ano}, para {self.get_descricao_servico()} {self.objeto}\n"
            f"Processo Administrativo NUP: {self.nup}\n"
            f"Setor Demandante: {self.setor_responsavel}"
        )
        self.update_text_fields()

    def handle_gerar_aviso_dispensa(self):
        self.assunto_text = f"{self.id_processo} - Aviso de Dispensa Eletrônica"
        self.sinopse_text = (
            f"Aviso referente à {self.tipo} nº {self.numero}/{self.ano}, para {self.get_descricao_servico()} {self.objeto}\n"
            f"Processo Administrativo NUP: {self.nup}\n"
            f"Setor Demandante: {self.setor_responsavel}"
        )
        self.update_text_fields()
        self.consolidador.gerar_aviso_dispensa()

    def handle_gerar_aviso_dispensa_sidgem(self):
        self.assunto_text = f"{self.id_processo} - Aviso de Dispensa Eletrônica"
        self.sinopse_text = (
            f"Aviso referente à {self.tipo} nº {self.numero}/{self.ano}, para {self.get_descricao_servico()} {self.objeto}\n"
            f"Processo Administrativo NUP: {self.nup}\n"
            f"Setor Demandante: {self.setor_responsavel}"
        )
        self.update_text_fields()

    def update_text_fields(self):
        self.textEditAssunto.setPlainText(self.assunto_text)
        self.textEditSinopse.setPlainText(self.sinopse_text)

    def create_GrupoSIGDEM(self):       
        grupoSIGDEM = QGroupBox("SIGDEM")
        self.apply_widget_style(grupoSIGDEM)

        layout = QVBoxLayout(grupoSIGDEM)

        labelAssunto = QLabel("No campo “Assunto”:")
        labelAssunto.setStyleSheet("font-size: 12pt;")
        layout.addWidget(labelAssunto)
        self.textEditAssunto = QTextEdit(f"{self.id_processo} - Abertura de Dispensa Eletrônica")
        self.textEditAssunto.setStyleSheet("font-size: 12pt;")
        self.textEditAssunto.setMaximumHeight(60)
        layoutHAssunto = QHBoxLayout()
        layoutHAssunto.addWidget(self.textEditAssunto)
        icon_copy = QIcon(str(self.ICONS_DIR / "copy_1.png"))
        btnCopyAssunto = self.create_button(text="", icon=icon_copy, callback=lambda: self.copyToClipboard(self.textEditAssunto.toPlainText()), tooltip_text="Copiar texto para a área de transferência", button_size=QSize(40, 40), icon_size=QSize(25, 25))
        layoutHAssunto.addWidget(btnCopyAssunto)
        layout.addLayout(layoutHAssunto)

        labelSinopse = QLabel("No campo “Sinopse”:")
        labelSinopse.setStyleSheet("font-size: 12pt;")
        layout.addWidget(labelSinopse)
        self.textEditSinopse = QTextEdit(
            f"Termo de Abertura referente à {self.tipo} nº {self.numero}/{self.ano}, para {self.get_descricao_servico()} {self.objeto}\n"
            f"Processo Administrativo NUP: {self.nup}\n"
            f"Setor Demandante: {self.setor_responsavel}"
        )
        self.textEditSinopse.setStyleSheet("font-size: 12pt;")
        self.textEditSinopse.setMaximumHeight(140)
        layoutHSinopse = QHBoxLayout()
        layoutHSinopse.addWidget(self.textEditSinopse)
        btnCopySinopse = self.create_button(text="", icon=icon_copy, callback=lambda: self.copyToClipboard(self.textEditSinopse.toPlainText()), tooltip_text="Copiar texto para a área de transferência", button_size=QSize(40, 40), icon_size=QSize(25, 25))
        layoutHSinopse.addWidget(btnCopySinopse)
        layout.addLayout(layoutHSinopse)

        grupoSIGDEM.setLayout(layout)
        self.carregarAgentesResponsaveis()
        
        return grupoSIGDEM

    def get_descricao_servico(self):
        return "aquisição de" if self.material_servico == "Material" else "contratação de empresa especializada em"

    def copyToClipboard(self, text):
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        QToolTip.showText(QCursor.pos(), "Texto copiado para a área de transferência.", msecShowTime=1500)

    def format_cp_edit(self):
        text = self.cp_edit.text().strip()
        if '-' not in text:
            if text.isdigit():
                text = f"30-{int(text):02d}"
        else:
            parts = text.split('-')
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                self.cp_edit.setText(f"{int(parts[0]):02d}-{int(parts[1]):02d}")
                return
        
        self.cp_edit.setText(text)

    def on_autorizacao_clicked(self):
        print("Botão Autorização clicado")  # Substitua esta função pela funcionalidade desejada

    def importar_tabela(self):
        pass

    def abrir_pasta(self, pasta):
        QDesktopServices.openUrl(QUrl.fromLocalFile(str(pasta)))

    def verificar_subpasta(self):
        id_processo_modificado = self.id_processo.replace("/", "-")
        objeto_modificado = self.objeto.replace("/", "-")
        pastas_encontradas = []
        for subpasta in self.pasta_base.iterdir():
            if subpasta.is_dir() and id_processo_modificado in subpasta.name and objeto_modificado in subpasta.name:
                pastas_encontradas.append(subpasta.name)
                print(f"Pasta encontrada: {subpasta.name}")
        return pastas_encontradas

    def verificar_arquivo_pdf(self, pasta):
        arquivos_pdf = []
        if not pasta.exists():
            print(f"Pasta não encontrada: {pasta}")
            return None
        for arquivo in pasta.iterdir():
            if arquivo.suffix.lower() == ".pdf":
                arquivos_pdf.append(arquivo)
                # print(f"Arquivo PDF encontrado: {arquivo.name}")
        if arquivos_pdf:
            return max(arquivos_pdf, key=lambda p: p.stat().st_mtime)  # Retorna o PDF mais recente
        return None
    
    def verificar_e_criar_pastas(self, pasta_base):
        try:
            id_processo_modificado = self.id_processo.replace("/", "-")
            objeto_modificado = self.objeto.replace("/", "-")
            base_path = pasta_base / f'{id_processo_modificado} - {objeto_modificado}'

            pastas_necessarias = [
                pasta_base / '1. Autorizacao',
                pasta_base / '2. CP e anexos',
                pasta_base / '3. Aviso',
                pasta_base / '2. CP e anexos' / 'DFD',
                pasta_base / '2. CP e anexos' / 'DFD' / 'Anexo A - Relatorio Safin',
                pasta_base / '2. CP e anexos' / 'DFD' / 'Anexo B - Especificações e Quantidade',
                pasta_base / '2. CP e anexos' / 'TR',
                pasta_base / '2. CP e anexos' / 'TR' / 'Pesquisa de Preços',
                pasta_base / '2. CP e anexos' / 'Declaracao de Adequação Orçamentária',
                pasta_base / '2. CP e anexos' / 'Declaracao de Adequação Orçamentária' / 'Relatório do PDM-Catser',
                pasta_base / '2. CP e anexos' / 'Justificativas Relevantes',
            ]

            for pasta in pastas_necessarias:
                if not pasta.exists():
                    pasta.mkdir(parents=True)

        except (FileNotFoundError, PermissionError) as e:
            QMessageBox.critical(self, "Erro ao criar pastas", f"Não foi possível criar as pastas necessárias devido ao erro: {str(e)}. Por favor, selecione uma nova pasta base na aba 'Documentos'.")
            
        return pastas_necessarias

    def abrirPasta(self):
        print("Abrir pasta")
        pass    

    def add_pdf_to_merger(self):
        cp_number = self.cp_edit.text()
        if cp_number:
            pastas_necessarias = self.verificar_e_criar_pastas(self.pasta_base)
            pdf_add_dialog = PDFAddDialog(self.df_registro_selecionado, self.ICONS_DIR, pastas_necessarias, self.pasta_base, self)
            if pdf_add_dialog.exec():
                print(f"Adicionar PDF para CP nº {cp_number}")
            else:
                print("Ação de adicionar PDF cancelada.")
        else:
            QMessageBox.warning(self, "Erro", "Por favor, insira um número de CP válido.")

    def on_cp_clicked(self):
        # Implementação do callback para o botão CP
        pass

    def atualizar_action(self):
        icon_confirm = QIcon(str(self.ICONS_DIR / "confirm_green.png"))
        icon_x = QIcon(str(self.ICONS_DIR / "cancel.png"))

        def atualizar_anexo(section_title, anexo, label):
            pasta_anexo = None
            id_processo_modificado = self.id_processo.replace("/", "-")
            objeto_modificado = self.objeto.replace("/", "-")

            if section_title == "Documento de Formalização de Demanda (DFD)":
                if "Anexo A" in anexo:
                    pasta_anexo = self.pasta_base / f'{id_processo_modificado} - {objeto_modificado}' / '2. CP e anexos' / 'DFD' / 'Anexo A - Relatorio Safin'
                elif "Anexo B" in anexo:
                    pasta_anexo = self.pasta_base / f'{id_processo_modificado} - {objeto_modificado}' / '2. CP e anexos' / 'DFD' / 'Anexo B - Especificações e Quantidade'
            elif section_title == "Termo de Referência (TR)":
                pasta_anexo = self.pasta_base / f'{id_processo_modificado} - {objeto_modificado}' / '2. CP e anexos' / 'TR' / 'Pesquisa de Preços'
            elif section_title == "Declaração de Adequação Orçamentária":
                pasta_anexo = self.pasta_base / f'{id_processo_modificado} - {objeto_modificado}' / '2. CP e anexos' / 'Declaracao de Adequação Orçamentária' / 'Relatório do PDM-Catser'

            if pasta_anexo:
                print(f"Verificando pasta: {pasta_anexo}")
                arquivos_pdf = self.verificar_arquivo_pdf(pasta_anexo)
                icon = icon_confirm if arquivos_pdf else icon_x
                label.setPixmap(icon.pixmap(QSize(30, 30)))
            else:
                print(f"Anexo não identificado: {anexo}")
                label.setPixmap(icon_x.pixmap(QSize(30, 30)))

        for section_title, anexos in self.anexos_dict.items():
            for anexo, icon_label in anexos:
                atualizar_anexo(section_title, anexo, icon_label)

        self.dados_atualizados.emit()

    def legenda_action(self):
        pass

    """
    
    
    
    



    
    PARTE REFERENTE AO GROUPO DO SETOR RESPONSÁVEL
    
    






    """

    def create_dados_responsavel_contratacao_group(self):
        data = self.extract_registro_data()

        setor_responsavel_group_box = QGroupBox("Divisão/Setor Responsável pela Demanda")
        self.apply_widget_style(setor_responsavel_group_box)
        setor_responsavel_layout = QVBoxLayout()

        # Configuração da OM, Divisão, e CP na mesma linha
        om_divisao_layout = QHBoxLayout()

        # Configuração da OM
        om_layout = QHBoxLayout()
        om_label = QLabel("OM:")
        self.apply_widget_style(om_label)
        
        sigla_om = data.get('sigla_om', 'CeIMBra')
        if self.df_registro_selecionado is not None and 'sigla_om' in self.df_registro_selecionado.columns:
            sigla_om = self.df_registro_selecionado['sigla_om'].iloc[0] if not self.df_registro_selecionado['sigla_om'].empty else 'CeIMBra'

        self.om_combo = self.create_combo_box(sigla_om, [], 150, 35)
        om_layout.addWidget(om_label)
        om_layout.addWidget(self.om_combo)

        # Adicionando o layout OM ao layout principal
        om_divisao_layout.addLayout(om_layout)

        # Configuração da Divisão
        divisao_label = QLabel("Divisão:")
        self.apply_widget_style(divisao_label)

        # Criando o QComboBox editável
        self.setor_responsavel_combo = QComboBox()
        self.setor_responsavel_combo.setEditable(True)

        # Adicionando as opções ao ComboBox
        divisoes = [
            "Divisão de Abastecimento",
            "Divisão de Finanças",
            "Divisão de Obtenção",
            "Divisão de Pagamento",
            "Divisão de Administração",
            "Divisão de Subsistência"
        ]
        self.setor_responsavel_combo.addItems(divisoes)

        # Definindo o texto atual com base nos dados fornecidos
        self.setor_responsavel_combo.setCurrentText(data['setor_responsavel'])

        # Definindo a política de tamanho para expandir e preencher o espaço disponível
        self.setor_responsavel_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Adicionando o QComboBox ao layout
        om_divisao_layout.addWidget(divisao_label)
        om_divisao_layout.addWidget(self.setor_responsavel_combo)

        # Adicionando o layout OM/Divisão/CP ao layout principal
        setor_responsavel_layout.addLayout(om_divisao_layout)
        
        self.load_sigla_om(sigla_om)  # Carregar os itens do combobox e definir o texto

        self.par_edit = QLineEdit(str(data.get('cod_par', '')))
        self.par_edit.setFixedWidth(150)
        self.prioridade_combo = self.create_combo_box(data.get('prioridade_par', 'Necessário'), ["Necessário", "Urgente", "Desejável"], 190, 35)
     
        
        par_layout = QHBoxLayout()

        # Configuração da CP
        cp_label = QLabel("Número da CP:")
        self.apply_widget_style(cp_label)
        self.cp_edit = QLineEdit(data['comunicacao_padronizada'])
        self.cp_edit.setFixedWidth(150)  # Ajuste do tamanho para 50
        par_layout.addWidget(cp_label)
        par_layout.addWidget(self.cp_edit)

        par_label = QLabel("Meta do PAR:")
        prioridade_label = QLabel("Prioridade:")
        self.apply_widget_style(par_label)
        self.apply_widget_style(prioridade_label)
        par_layout.addWidget(par_label)
        par_layout.addWidget(self.par_edit)
        par_layout.addWidget(prioridade_label)
        par_layout.addWidget(self.prioridade_combo)
        setor_responsavel_layout.addLayout(par_layout)

        self.endereco_edit = QLineEdit(data['endereco'])
        self.endereco_edit.setFixedWidth(450)
        self.cep_edit = QLineEdit(str(data.get('cep', '')))
        endereco_cep_layout = QHBoxLayout()
        endereco_label = QLabel("Endereço:")
        cep_label = QLabel("CEP:")
        self.apply_widget_style(endereco_label)
        self.apply_widget_style(cep_label)
        endereco_cep_layout.addWidget(endereco_label)
        endereco_cep_layout.addWidget(self.endereco_edit)
        endereco_cep_layout.addWidget(cep_label)
        endereco_cep_layout.addWidget(self.cep_edit)
        setor_responsavel_layout.addLayout(endereco_cep_layout)

        self.email_edit = QLineEdit(data['email'])
        self.email_edit.setFixedWidth(400)
        self.telefone_edit = QLineEdit(data['telefone'])
        email_telefone_layout = QHBoxLayout()
        email_telefone_layout.addLayout(self.create_layout("E-mail:", self.email_edit))
        email_telefone_layout.addLayout(self.create_layout("Tel:", self.telefone_edit))
        setor_responsavel_layout.addLayout(email_telefone_layout)

        self.dias_edit = QLineEdit("Segunda à Sexta")
        setor_responsavel_layout.addLayout(self.create_layout("Dias para Recebimento:", self.dias_edit))

        self.horario_edit = QLineEdit("09 às 11h20 e 14 às 16h30")
        setor_responsavel_layout.addLayout(self.create_layout("Horário para Recebimento:", self.horario_edit))

        # Adicionando Justificativa
        justificativa_label = QLabel("Justificativa para a contratação:")
        justificativa_label.setStyleSheet("font-size: 12pt;")
        self.justificativa_edit = QTextEdit(self.get_justification_text())
        self.apply_widget_style(self.justificativa_edit)
        setor_responsavel_layout.addWidget(justificativa_label)
        setor_responsavel_layout.addWidget(self.justificativa_edit)
        # setor_responsavel_layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))
        setor_responsavel_group_box.setLayout(setor_responsavel_layout)

        return setor_responsavel_group_box

    def get_justification_text(self):
        # Recupera o valor atual da justificativa no DataFrame
        current_justification = self.df_registro_selecionado['justificativa'].iloc[0]

        # Retorna o valor atual se ele existir, senão, constrói uma justificativa baseada no tipo de material/serviço
        if current_justification:  # Checa se existe uma justificativa
            return current_justification
        else:
            # Gera justificativa padrão com base no tipo de material ou serviço
            if self.material_servico == 'Material':
                return (f"A aquisição de {self.objeto} se faz necessária para o atendimento das necessidades do(a) {self.setor_responsavel} do(a) {self.orgao_responsavel} ({self.sigla_om}). A disponibilidade e a qualidade dos materiais são essenciais para garantir a continuidade das operações e a eficiência das atividades desempenhadas pelo(a) {self.setor_responsavel}.")
            elif self.material_servico == 'Serviço':
                return (f"A contratação de empresa especializada na prestação de serviços de {self.objeto} é imprescindível para o atendimento das necessidades do(a) {self.setor_responsavel} do(a) {self.orgao_responsavel} ({self.sigla_om}).")
            return ""  # Retorna uma string vazia se nenhuma condição acima for satisfeita

    def load_sigla_om(self, sigla_om):
        try:
            with sqlite3.connect(self.database_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT sigla_om FROM controle_om ORDER BY sigla_om")
                items = [row[0] for row in cursor.fetchall()]
                self.om_combo.addItems(items)
                self.om_combo.setCurrentText(sigla_om)  # Define o texto atual do combobox
                self.om_combo.currentTextChanged.connect(self.on_om_changed)
                print(f"Loaded sigla_om items: {items}")
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao carregar OM: {e}")
            print(f"Error loading sigla_om: {e}")

    def on_om_changed(self):
        selected_om = self.om_combo.currentText()
        print(f"OM changed to: {selected_om}")
        with sqlite3.connect(self.database_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT uasg, orgao_responsavel FROM controle_om WHERE sigla_om = ?", (selected_om,))
            result = cursor.fetchone()
            if result:
                uasg, orgao_responsavel = result
                index = self.df_registro_selecionado.index[0]
                self.df_registro_selecionado.loc[index, 'uasg'] = uasg
                self.df_registro_selecionado.loc[index, 'orgao_responsavel'] = orgao_responsavel
                print(f"Updated DataFrame: uasg={uasg}, orgao_responsavel={orgao_responsavel}")
                self.title_updated.emit(f"{orgao_responsavel} (UASG: {uasg})")

class FormularioExcel:
    def __init__(self, df_registro_selecionado, pasta_base, parent_dialog):
        self.df_registro_selecionado = df_registro_selecionado
        self.pasta_base = Path(pasta_base)
        self.parent_dialog = parent_dialog

        self.colunas_legiveis = {
            'nup': 'NUP',
            'material_servico': 'Material (M) ou Serviço (S)',
            'objeto': 'Objeto Resumido',
            'vigencia': 'Vigência',
            'criterio_julgamento': 'Critério de Julgamento (Menor Preço ou Maior Desconto)',
            'com_disputa': 'Com disputa? Sim (S) ou Não (N)',
            'pesquisa_preco': 'Pesquisa Concomitante? Sim (S) ou Não (N)',
            'previsao_contratacao': 'Previsão de Contratação',
            'uasg': 'UASG',
            'setor_responsavel': 'Setor Responsável',
            'cod_par': 'Código PAR',
            'prioridade_par': 'Prioridade PAR (Necessário, Urgente ou Desejável)',
            'cep': 'CEP',
            'endereco': 'Endereço',
            'email': 'Email',
            'telefone': 'Telefone',
            'dias_para_recebimento': 'Dias para Recebimento',
            'horario_para_recebimento': 'Horário para Recebimento',
            'valor_total': 'Valor Total',
            'acao_interna': 'Ação Interna',
            'fonte_recursos': 'Fonte de Recursos',
            'natureza_despesa': 'Natureza da Despesa',
            'unidade_orcamentaria': 'Unidade Orçamentária',
            'programa_trabalho_resuminho': 'PTRES',
            'atividade_custeio': 'Atividade de Custeio',
            'justificativa': 'Justificativa',
            'comunicacao_padronizada': 'Comunicação Padronizada (CP), Ex: 60-25',
            'do_responsavel': 'Campo Do(a) da CP',
            'ao_responsavel': 'Campo Ao da CP'
        }

        self.normalizacao_valores = {
            'material_servico': {
                'M': 'Material',
                'm': 'Material',
                'Material': 'Material',
                'material': 'Material',
                'S': 'Serviço',
                's': 'Serviço',
                'Serviço': 'Serviço',
                'serviço': 'Serviço',
                'Servico': 'Serviço',
                'servico': 'Serviço'
            },
            'com_disputa': {
                'S': 'Sim',
                's': 'Sim',
                'Sim': 'Sim',
                'sim': 'Sim',
                'N': 'Não',
                'n': 'Não',
                'Não': 'Não',
                'não': 'Não',
                'Nao': 'Não',
                'nao': 'Não'
            },
            'pesquisa_preco': {
                'S': 'Sim',
                's': 'Sim',
                'Sim': 'Sim',
                'sim': 'Sim',
                'N': 'Não',
                'n': 'Não',
                'Não': 'Não',
                'não': 'Não',
                'Nao': 'Não',
                'nao': 'Não'
            },
            'atividade_custeio': {
                'S': 'Sim',
                's': 'Sim',
                'Sim': 'Sim',
                'sim': 'Sim',
                'N': 'Não',
                'n': 'Não',
                'Não': 'Não',
                'não': 'Não',
                'Nao': 'Não',
                'nao': 'Não'
            }
        }


        self.colunas_legiveis_inverso = {v: k for k, v in self.colunas_legiveis.items()}


    def criar_formulario(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Formulário"

            df_filtrado = self._filtrar_dataframe()
            self._adicionar_titulo(ws)
            self._definir_cabecalhos(ws)
            self._preencher_dados(ws, df_filtrado)
            self._aplicar_bordas(ws)
            
            file_path = self._salvar_arquivo(wb)
            self._abrir_arquivo(file_path)

            QMessageBox.information(None, "Sucesso", "Formulário criado e aberto com sucesso.")
        except Exception as e:
            print(f"Erro ao criar formulário: {str(e)}")
            QMessageBox.critical(None, "Erro", f"Falha ao criar formulário: {str(e)}")

    def carregar_formulario(self):
        try:
            print("DataFrame antes de carregar o formulário:")
            print(self.df_registro_selecionado)

            file_path, _ = QFileDialog.getOpenFileName(None, "Selecione o formulário", "", "Excel Files (*.xlsx *.ods);;All Files (*)")
            if not file_path:
                return

            if file_path.endswith('.xlsx'):
                wb = load_workbook(file_path)
                ws = wb.active

                if ws['A2'].value != "Índice" or ws['B2'].value != "Valor":
                    QMessageBox.critical(None, "Erro", "O formulário selecionado está incorreto.")
                    return

                for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
                    coluna_legivel = row[0]
                    valor = row[1]
                    coluna = self.colunas_legiveis_inverso.get(coluna_legivel, coluna_legivel)
                    if coluna in self.normalizacao_valores:
                        valor = self.normalizacao_valores[coluna].get(valor, valor)
                    if coluna in self.df_registro_selecionado.columns:
                        self.df_registro_selecionado.at[0, coluna] = valor

            elif file_path.endswith('.ods'):
                df = pd.read_excel(file_path, engine='odf')

                if df.iloc[0, 0] != "Índice" or df.iloc[0, 1] != "Valor":
                    QMessageBox.critical(None, "Erro", "O formulário selecionado está incorreto.")
                    return

                for _, row in df.iloc[1:].iterrows():
                    coluna_legivel = row[0]
                    valor = row[1]
                    coluna = self.colunas_legiveis_inverso.get(coluna_legivel, coluna_legivel)
                    if coluna in self.normalizacao_valores:
                        valor = self.normalizacao_valores[coluna].get(valor, valor)
                    if coluna in self.df_registro_selecionado.columns:
                        self.df_registro_selecionado.at[0, coluna] = valor


            print("DataFrame após carregar o formulário:")
            print(self.df_registro_selecionado)

            self.parent_dialog.preencher_campos()
            self.parent_dialog.dados_atualizados.emit()

            QMessageBox.information(None, "Sucesso", "Formulário carregado com sucesso.")
        except Exception as e:
            print(f"Erro ao carregar formulário: {str(e)}")
            QMessageBox.critical(None, "Erro", f"Falha ao carregar formulário: {str(e)}")


    def _filtrar_dataframe(self):
        colunas_incluir = list(self.colunas_legiveis.keys())
        df_filtrado = self.df_registro_selecionado[colunas_incluir].rename(columns=self.colunas_legiveis)
        return df_filtrado

    def _adicionar_titulo(self, ws):
        tipo = self.df_registro_selecionado['tipo'].iloc[0]
        numero = self.df_registro_selecionado['numero'].iloc[0]
        ano = self.df_registro_selecionado['ano'].iloc[0]
        titulo = f"{tipo} nº {numero}/{ano}"
        ws.merge_cells('A1:B1')
        ws['A1'] = titulo
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40

    def _definir_cabecalhos(self, ws):
        ws['A2'] = "Índice"
        ws['B2'] = "Valor"
        ws['A2'].font = Font(size=14, bold=True)
        ws['B2'].font = Font(size=14, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws['A2'].border = thin_border
        ws['B2'].border = thin_border
        ws.column_dimensions[get_column_letter(1)].width = 50
        ws.column_dimensions[get_column_letter(2)].width = 80

    def _preencher_dados(self, ws, df_filtrado):
        for i, (col_name, value) in enumerate(df_filtrado.iloc[0].items(), start=3):
            ws[f'A{i}'] = col_name
            ws[f'B{i}'] = str(value)
            ws[f'B{i}'].alignment = Alignment(wrap_text=True)
            fill_color = "F2F2F2" if i % 2 == 0 else "FFFFFF"
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws[f'A{i}'].fill = fill
            ws[f'B{i}'].fill = fill
            ws[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[i].height = 60 if i == 28 else 15

    def _aplicar_bordas(self, ws):
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

    def _salvar_arquivo(self, wb):
        file_path = self.pasta_base / "formulario.xlsx"
        wb.save(file_path)
        return file_path

    def _abrir_arquivo(self, file_path):
        if os.name == 'nt':
            os.startfile(file_path)
        elif os.name == 'posix':
            subprocess.call(['open', file_path])
        else:
            subprocess.call(['xdg-open', file_path])

class StackWidgetDispensaManager:
    def __init__(self, parent, default_widget=None):
        self.stack = QStackedWidget(parent)
        self.widgets = {}
        if default_widget:
            self.add_widget("default", default_widget)

    def add_widget(self, name, widget):
        self.widgets[name] = widget
        self.stack.addWidget(widget)

    def show_widget(self, name):
        widget = self.widgets.get(name)
        if widget:
            self.stack.setCurrentWidget(widget)
        else:
            print(f"Widget '{name}' não encontrado.")

    def get_widget(self):
        return self.stack   

    def fill_frame_dados_secundarios(self):
        data = self.extract_registro_data()
        detalhes_layout = QVBoxLayout()

        hbox_top_layout = QHBoxLayout()  # Layout horizontal para os três QGroupBox

        # Preenche os QGroupBox e os adiciona ao layout horizontalhttps://academico.movelms.com/ExamReference/GetExam/17179682#
        contratacao_group_box = self.create_contratacao_group(data)        
        dados_do_setor_responsavel_contratacao_group_box = self.fill_frame_dados_do_setor_resposavel_contratacao()
        sigdem_group = self.setupGrupoSIGDEM()
        agentes_responsaveis_group = self.fill_frame_agentes_responsaveis()

        hbox_top_layout.addWidget(contratacao_group_box)
        hbox_top_layout.addWidget(dados_do_setor_responsavel_contratacao_group_box)
        hbox_top_layout.addWidget(sigdem_group)
        hbox_top_layout.addWidget(agentes_responsaveis_group)

        # Adiciona o layout horizontal ao layout principal
        detalhes_layout.addLayout(hbox_top_layout)

        hbox_down_layout = QHBoxLayout()  # Layout horizontal para os QGroupBox
        # Preenche os QGroupBox e os adiciona ao layout horizontal
        classificacao_orcamentaria_group_box = self.fill_frame_classificacao_orcamentaria()
        comunicacao_padronizada_group = self.fill_frame_comunicacao_padronizada()
        # lista_verificacao_group = self.fill_frame_criar_documentos()
        formulario_group = self.fill_frame_formulario()
        
        # Adiciona o gerar_documentos_group_box e utilidades_group em um layout vertical
        vertical_widget = QWidget()
        vertical_layout = QVBoxLayout()
        vertical_layout.setContentsMargins(0, 0, 0, 0) 
        vertical_layout.setSpacing(0)
        vertical_widget.setLayout(vertical_layout)
        
        gerar_documentos_group_box = self.create_gerar_documentos_group()
        utilidades_group = self.fill_frame_utilidades()

        vertical_layout.addWidget(gerar_documentos_group_box)
        vertical_layout.addWidget(utilidades_group)

        # Criação e configuração da label de imagem fora do grupo de formulário
        caminho_imagem = IMAGE_PATH / "licitacao_360.png" 
        licitacao_360_pixmap = QPixmap(str(caminho_imagem))
        licitacao_360_pixmap = licitacao_360_pixmap.scaled(240, 240, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)

        image_label = QLabel()
        image_label.setPixmap(licitacao_360_pixmap)
        image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Criação de um widget vertical para agrupar elementos
        vertical_widget_formulario_image = QWidget()
        vertical_layout_formulario_image = QVBoxLayout()
        vertical_layout_formulario_image.setContentsMargins(0, 0, 0, 0) 
        vertical_widget_formulario_image.setLayout(vertical_layout_formulario_image)

        vertical_layout_formulario_image.addWidget(self.fill_frame_formulario())
        vertical_layout_formulario_image.addWidget(image_label)

        hbox_down_layout.addWidget(classificacao_orcamentaria_group_box)
        hbox_down_layout.addWidget(comunicacao_padronizada_group)
        
        # hbox_down_layout.addWidget(lista_verificacao_group)
        hbox_down_layout.addWidget(vertical_widget)  # Adiciona o QWidget ao layout horizontal
        hbox_down_layout.addWidget(vertical_widget_formulario_image)

        # Adiciona o layout horizontal ao layout principal
        detalhes_layout.addLayout(hbox_down_layout)
        detalhes_layout.setContentsMargins(0, 0, 0, 0) 
        detalhes_layout.setSpacing(0)

        self.frame_secundario_layout.addLayout(detalhes_layout)
    
class StackWidgetManager:
    def __init__(self, parent, data_function):
        """
        Inicializa o gerenciador de StackWidget.
        
        Args:
            parent: O widget pai onde o QStackedWidget será inserido.
            data_function: Função que retorna os dados para criação dos widgets.
        """
        self.parent = parent
        self.data_function = data_function
        self.stacked_widget = QStackedWidget(parent)
        self.stacked_widget.setStyleSheet(
            "QStackedWidget {"
            "border: 1px solid #414242; border-radius: 5px; "
            "border-bottom-left-radius: 5px; border-bottom-right-radius: 5px; }"
        )
        self.widgets = {}  # Armazena os widgets adicionados com seus nomes
        self.widget_creators = {}  # Funções para criar os widgets

    def add_widget(self, name, widget):
        """
        Adiciona um widget ao QStackedWidget e mapeia pelo nome.
        
        Args:
            name (str): Nome do widget para referência futura.
            widget (QWidget): Instância do widget a ser adicionado.
        """
        self.stacked_widget.addWidget(widget)
        self.widgets[name] = widget

    def show_widget(self, name):
        """
        Mostra o widget especificado pelo nome.
        
        Args:
            name (str): Nome do widget a ser mostrado.
        """
        if name in self.widgets:
            self.stacked_widget.setCurrentWidget(self.widgets[name])
        else:
            create_widget_function = self.widget_creators.get(name)
            if create_widget_function:
                # Cria e mostra o widget, caso ainda não tenha sido criado
                self.add_widget(name, create_widget_function())
                self.stacked_widget.setCurrentWidget(self.widgets[name])

    def register_widget_creator(self, name, creator_function):
        """
        Registra uma função que cria um widget, associada a um nome.
        
        Args:
            name (str): Nome do widget.
            creator_function (function): Função que retorna uma instância do widget.
        """
        self.widget_creators[name] = creator_function
