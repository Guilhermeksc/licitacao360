from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import *
from diretorios import *
from pathlib import Path
import pandas as pd
import os
import subprocess
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class FormularioExcel:
    def __init__(self, df_registro_selecionado, pasta_base, parent_dialog):
        self.df_registro_selecionado = df_registro_selecionado
        self.pasta_base = Path(pasta_base)
        self.parent_dialog = parent_dialog

        self.colunas_legiveis = {
            'nup': 'NUP',
            'material_servico': 'Material (M) ou Serviço (S)',
            'vigencia': 'Vigência',
            'criterio_julgamento': 'Critério de Julgamento (Menor Preço ou Maior Desconto)',
            'com_disputa': 'Com disputa? Sim (S) ou Não (N)',
            'pesquisa_preco': 'Pesquisa Concomitante? Sim (S) ou Não (N)',
            'previsao_contratacao': 'Previsão de Contratação',
            'uasg': 'UASG',
            'setor_responsavel': 'Setor Responsável',
            'cod_par': 'Código PAR',
            'prioridade_par': 'Prioridade PAR (Necessário, Urgente ou Desejável)',
            'cep': 'CEP',
            'endereco': 'Endereço',
            'email': 'Email',
            'telefone': 'Telefone',
            'dias_para_recebimento': 'Dias para Recebimento',
            'horario_para_recebimento': 'Horário para Recebimento',
            'valor_total': 'Valor Total',
            'acao_interna': 'Ação Interna',
            'fonte_recursos': 'Fonte de Recursos',
            'natureza_despesa': 'Natureza da Despesa',
            'unidade_orcamentaria': 'Unidade Orçamentária',
            'ptres': 'PTRES',
            'atividade_custeio': 'Atividade de Custeio',
            'justificativa': 'Justificativa',
            'comunicacao_padronizada': 'Comunicação Padronizada (CP), Ex: 60-25',
        }

        self.normalizacao_valores = {
            'material_servico': {
                'M': 'Material',
                'm': 'Material',
                'Material': 'Material',
                'material': 'Material',
                'S': 'Serviço',
                's': 'Serviço',
                'Serviço': 'Serviço',
                'serviço': 'Serviço',
                'Servico': 'Serviço',
                'servico': 'Serviço'
            },
            'com_disputa': {
                'S': 'Sim',
                's': 'Sim',
                'Sim': 'Sim',
                'sim': 'Sim',
                'N': 'Não',
                'n': 'Não',
                'Não': 'Não',
                'não': 'Não',
                'Nao': 'Não',
                'nao': 'Não'
            },
            'pesquisa_preco': {
                'S': 'Sim',
                's': 'Sim',
                'Sim': 'Sim',
                'sim': 'Sim',
                'N': 'Não',
                'n': 'Não',
                'Não': 'Não',
                'não': 'Não',
                'Nao': 'Não',
                'nao': 'Não'
            },
            'atividade_custeio': {
                'S': 'Sim',
                's': 'Sim',
                'Sim': 'Sim',
                'sim': 'Sim',
                'N': 'Não',
                'n': 'Não',
                'Não': 'Não',
                'não': 'Não',
                'Nao': 'Não',
                'nao': 'Não'
            }
        }


        self.colunas_legiveis_inverso = {v: k for k, v in self.colunas_legiveis.items()}


    def criar_formulario(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Formulário"

            df_filtrado = self._filtrar_dataframe()
            self._adicionar_titulo(ws)
            self._definir_cabecalhos(ws)
            self._preencher_dados(ws, df_filtrado)
            self._aplicar_bordas(ws)
            
            file_path = self._salvar_arquivo(wb)
            self._abrir_arquivo(file_path)

            QMessageBox.information(None, "Sucesso", "Formulário criado e aberto com sucesso.")
        except Exception as e:
            print(f"Erro ao criar formulário: {str(e)}")
            QMessageBox.critical(None, "Erro", f"Falha ao criar formulário: {str(e)}")

    def carregar_formulario(self):
        try:
            print("DataFrame antes de carregar o formulário:")
            print(self.df_registro_selecionado)

            file_path, _ = QFileDialog.getOpenFileName(None, "Selecione o formulário", "", "Excel Files (*.xlsx *.ods);;All Files (*)")
            if not file_path:
                return

            if file_path.endswith('.xlsx'):
                wb = load_workbook(file_path)
                ws = wb.active

                if ws['A2'].value != "Índice" or ws['B2'].value != "Valor":
                    QMessageBox.critical(None, "Erro", "O formulário selecionado está incorreto.")
                    return

                for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
                    coluna_legivel = row[0]
                    valor = row[1]
                    coluna = self.colunas_legiveis_inverso.get(coluna_legivel, coluna_legivel)
                    if coluna in self.normalizacao_valores:
                        valor = self.normalizacao_valores[coluna].get(valor, valor)
                    if coluna in self.df_registro_selecionado.columns:
                        self.df_registro_selecionado.at[0, coluna] = valor

            elif file_path.endswith('.ods'):
                df = pd.read_excel(file_path, engine='odf')

                if df.iloc[0, 0] != "Índice" or df.iloc[0, 1] != "Valor":
                    QMessageBox.critical(None, "Erro", "O formulário selecionado está incorreto.")
                    return

                for _, row in df.iloc[1:].iterrows():
                    coluna_legivel = row[0]
                    valor = row[1]
                    coluna = self.colunas_legiveis_inverso.get(coluna_legivel, coluna_legivel)
                    if coluna in self.normalizacao_valores:
                        valor = self.normalizacao_valores[coluna].get(valor, valor)
                    if coluna in self.df_registro_selecionado.columns:
                        self.df_registro_selecionado.at[0, coluna] = valor


            print("DataFrame após carregar o formulário:")
            print(self.df_registro_selecionado)

            self.parent_dialog.preencher_campos()
            self.parent_dialog.dados_atualizados.emit()

            QMessageBox.information(None, "Sucesso", "Formulário carregado com sucesso.")
        except Exception as e:
            print(f"Erro ao carregar formulário: {str(e)}")
            QMessageBox.critical(None, "Erro", f"Falha ao carregar formulário: {str(e)}")


    def _filtrar_dataframe(self):
        colunas_incluir = list(self.colunas_legiveis.keys())
        df_filtrado = self.df_registro_selecionado[colunas_incluir].rename(columns=self.colunas_legiveis)
        return df_filtrado

    def _adicionar_titulo(self, ws):
        tipo = self.df_registro_selecionado['tipo'].iloc[0]
        numero = self.df_registro_selecionado['numero'].iloc[0]
        ano = self.df_registro_selecionado['ano'].iloc[0]
        objeto = self.df_registro_selecionado['objeto'].iloc[0]
        titulo = f"{tipo} nº {numero}/{ano} ({objeto})"
        ws.merge_cells('A1:B1')
        ws['A1'] = titulo
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40

    def _definir_cabecalhos(self, ws):
        ws['A2'] = "Índice"
        ws['B2'] = "Valor"
        ws['A2'].font = Font(size=14, bold=True)
        ws['B2'].font = Font(size=14, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws['A2'].border = thin_border
        ws['B2'].border = thin_border
        ws.column_dimensions[get_column_letter(1)].width = 50
        ws.column_dimensions[get_column_letter(2)].width = 80

    def _preencher_dados(self, ws, df_filtrado):
        for i, (col_name, value) in enumerate(df_filtrado.iloc[0].items(), start=3):
            ws[f'A{i}'] = col_name
            ws[f'B{i}'] = str(value)
            ws[f'B{i}'].alignment = Alignment(wrap_text=True)
            fill_color = "F2F2F2" if i % 2 == 0 else "FFFFFF"
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws[f'A{i}'].fill = fill
            ws[f'B{i}'].fill = fill
            ws[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[i].height = 60 if i == 28 else 15

    def _aplicar_bordas(self, ws):
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

    def _salvar_arquivo(self, wb):
        file_path = self.pasta_base / "formulario.xlsx"
        wb.save(file_path)
        return file_path

    def _abrir_arquivo(self, file_path):
        if os.name == 'nt':
            os.startfile(file_path)
        elif os.name == 'posix':
            subprocess.call(['open', file_path])
        else:
            subprocess.call(['xdg-open', file_path])