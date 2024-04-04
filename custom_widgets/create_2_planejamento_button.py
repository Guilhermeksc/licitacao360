#create_2_planejamento_button.py

from PyQt6.QtWidgets import *
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QPixmap, QIcon, QFont, QMovie
from PyQt6.QtCore import *
from pathlib import Path
from diretorios import *
from utils.treeview_utils import load_images, create_button, save_dataframe_to_excel
from utils.utilidades import ler_arquivo_json, escrever_arquivo_json, inicializar_json_do_excel, sincronizar_json_com_dataframe
import pandas as pd
import subprocess
import win32com.client
import tempfile
import time
import json
import os
import sys
import datetime
from datetime import datetime
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import fitz

df_uasg = pd.read_excel(TABELA_UASG_DIR)
global df_registro_selecionado
df_registro_selecionado = None

class ReportButton(QPushButton):
    openReportDialog = pyqtSignal()

    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.clicked.connect(self.emitOpenReportDialogSignal)

    def emitOpenReportDialogSignal(self):
        self.openReportDialog.emit()

def status_sort_key(status):
    order = [
        'Concluído', 'Assinatura Contrato', 'Homologado', 'Em recurso',
        'Sessão Pública', 'Impugnado', 'Provisionamento', 'Recomendações AGU',
        'CJACM', 'Nota Técnica', 'Edital', 'IRP', 'Setor Responsável', 'Planejamento'
    ]
    try:
        return order.index(status)
    except ValueError:
        return len(order)
    
class ReportDialog(QDialog):
    def __init__(self, dataframe, icons_dir, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Relatório")
        self.setMinimumSize(980, 500)  # Define o tamanho mínimo do diálogo
        self.setLayout(QVBoxLayout())
        self.table_view = QTableView()
        self.layout().addWidget(self.table_view)
        self.model = QStandardItemModel()
        self.table_view.setModel(self.model)
        self.table_view.setStyleSheet("""
            QTableView {
                background-color: black;     
                color: white;
            }
        """)
        self.dataframe = dataframe  # Armazena o DataFrame passado como argumento
        self.icons_dir = Path(icons_dir)
        self.image_cache = {}
        self.image_cache = load_images(self.icons_dir, [
            "pdf64.png", "excel.png"
        ])
        # Configura os cabeçalhos das colunas
        self.model.setHorizontalHeaderLabels(["Número", "Objeto", "OM", "Status Anterior", "Dias", "Status Atual", "Dias", "Pregoeiro"])

        self.load_data()
        self._create_buttons()  # Cria os botões

    def showEvent(self, event):
        super().showEvent(event)
        self.adjust_column_widths()

    def adjust_column_widths(self):
        # Larguras fixas para as colunas conforme especificado
        column_widths = [75, 220, 80, 140, 30, 140, 30, 100]

        for column, width in enumerate(column_widths):
            self.table_view.setColumnWidth(column, width)

    def load_data(self):
        # Ler os dados do JSON
        try:
            with open(PROCESSOS_JSON_PATH, 'r', encoding='utf-8') as file:
                processos_json = json.load(file)
        except FileNotFoundError:
            print(f"Arquivo não encontrado: {PROCESSOS_JSON_PATH}")
            processos_json = {}

        # Ordena o DataFrame pelo 'Status Atual' usando a função de mapeamento
        self.dataframe['SortKey'] = self.dataframe['etapa'].apply(status_sort_key)
        self.dataframe.sort_values('SortKey', inplace=True)
        self.dataframe.drop('SortKey', axis=1, inplace=True)  # Remove a coluna auxiliar de ordenação

        for _, row in self.dataframe.iterrows():
            chave_processo = f"{row['mod']} {int(row['num_pregao'])}/{int(row['ano_pregao'])}"
            chave_processo_formatado = f"{row['mod']} {str(int(row['num_pregao'])).zfill(2)}/{int(row['ano_pregao'])}"
            processo = processos_json.get(chave_processo, {})
            historico = processo.get('historico', [])
            
            # Obter Status Anterior, Dias Status Anterior e Dias Status Atual
            status_anterior = historico[-2]['etapa'] if len(historico) >= 2 else '-'
            dias_status_anterior = str(historico[-2]['dias_na_etapa']) if len(historico) >= 2 else '-'
            dias_status_atual = str(historico[-1]['dias_na_etapa']) if historico else '-'

            self.model.appendRow([
                QStandardItem(chave_processo_formatado),
                QStandardItem(str(row['objeto']) if not pd.isna(row['objeto']) else ""),
                QStandardItem(str(row['sigla_om']) if not pd.isna(row['sigla_om']) else ""),
                QStandardItem(status_anterior),
                QStandardItem(dias_status_anterior),
                QStandardItem(str(row['etapa']) if not pd.isna(row['etapa']) else ""),
                QStandardItem(dias_status_atual),
                QStandardItem(str(row['pregoeiro']) if not pd.isna(row['pregoeiro']) else ""),
            ])
        QTimer.singleShot(10, self.adjust_column_widths)  # 100 ms após a UI ser mostrada

    def _create_buttons(self):
        # Cria um layout horizontal para os botões
        buttons_layout = QHBoxLayout()
        self.layout().addLayout(buttons_layout)  # Adiciona o layout de botões ao layout principal do diálogo

        # Especificações dos botões
        button_specs = [
            ("Tabela Excel", self.image_cache['excel'], self.on_export_excel, "Exportar dados para Excel"),
            ("Relatório PDF", self.image_cache['pdf64'], self.on_export_pdf, "Exportar dados para PDF")
        ]

        # Iterar sobre as especificações dos botões e criar cada botão
        for text, icon, callback, tooltip in button_specs:
            btn = create_button(text=text, icon=icon, callback=callback, tooltip_text=tooltip, parent=self)
            buttons_layout.addWidget(btn)  # Adiciona o botão ao layout de botões

    def create_excel(self, filename="relatorio.xlsx"):
        """
        Cria um arquivo Excel a partir dos dados do modelo, incluindo cabeçalhos personalizados e formatação.
        """
        # Cria um DataFrame dos dados
        data = []
        for row in range(self.model.rowCount()):
            row_data = []
            for column in range(self.model.columnCount()):
                item = self.model.item(row, column)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        df = pd.DataFrame(data, columns=[self.model.horizontalHeaderItem(i).text() for i in range(self.model.columnCount())])
        
        # Cria o arquivo Excel com XlsxWriter
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', startrow=4, index=False)  # A tabela começa na linha 5
        
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        # Configurações do formato de página e margens
        worksheet.set_landscape()  # Define o layout de página para paisagem
        worksheet.set_margins(left=0.79, right=0.39, top=0.39, bottom=0.39)  # Margens em polegadas (1 cm ≈ 0.39 inches, 2 cm ≈ 0.79 inches)
        worksheet.set_header('', options={'margin': 0})  # Cabeçalho com margem 0
        worksheet.set_footer('', options={'margin': 0})  # Rodapé com margem 0
                
        # Formatos para as células
        cabecalho_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'bold': True,
            'font_size': 14
        })
        cabecalho2_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'italic': True, 
            'font_size': 12
        })

        date_format = workbook.add_format({
            'italic': True, 
            'font_size': 10,
            'align': 'right'
        })

        # Formatos com cores intercaladas
        light_gray_format = workbook.add_format({'bg_color': '#F2F2F2', 'align': 'center', 'valign': 'vcenter'})
        white_format = workbook.add_format({'bg_color': '#FFFFFF', 'align': 'center', 'valign': 'vcenter'})
                
        # Configurações do cabeçalho e data
        worksheet.merge_range('A1:H1', 'Centro de Intendência da Marinha em Brasília', cabecalho_format)
        worksheet.merge_range('A2:H2', '"Prontidão e Efetividade no Planalto Central"', cabecalho2_format)
        worksheet.merge_range('A3:H3', 'Controle do Plano de Contratações Anual (PCA) 2024', cabecalho_format)
        data_atual = datetime.now().strftime("%d/%m/%Y")
        worksheet.merge_range('A4:H4', f"Atualizado em: {data_atual}", date_format)
        
        # Configurações de altura das linhas para o cabeçalho
        worksheet.set_row(0, 20)
        worksheet.set_row(2, 30)
        worksheet.set_row(3, 20)  # Ajuste de altura para a linha da data
            # Ajustar a largura das colunas, considerando a nova coluna 'Nº'
        col_widths = [10, 30, 10, 20, 5, 20, 5, 15]
        for i, width in enumerate(col_widths):
            worksheet.set_column(i, i, width)
        # Aplicar formatação de conteúdo centralizado a partir da linha 5
        for row_num in range(5, 5 + len(df)):
            for col_num in range(8):  # Colunas A a H
                cell_format = light_gray_format if (row_num % 2 == 0) else white_format
                worksheet.write(row_num, col_num, df.iloc[row_num - 5, col_num], cell_format)
        
        # Fecha o arquivo Excel
        writer.close()
        return filename  # Retorna o nome do arquivo criado

    def open_excel_file(self, filename):
        """
        Abre um arquivo Excel específico, usando um comando adequado dependendo do sistema operacional.
        """
        if os.name == 'nt':  # Para Windows
            os.startfile(filename)
        else:  # Para macOS e Linux
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, filename])

    def on_export_excel(self):
        filename = self.create_excel()  # Cria o arquivo Excel
        self.open_excel_file(filename)  # Abre o arquivo Excel criado

    def excel_to_pdf(self, excel_file_path, pdf_file_path):
        """
        Converte um arquivo Excel em PDF.
        """
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Executa em background
        try:
            doc = excel.Workbooks.Open(excel_file_path)
            doc.ExportAsFixedFormat(0, pdf_file_path)  # 0 indica que estamos exportando para PDF
        except Exception as e:
            print(f"Erro: {e}")
        finally:
            if doc is not None:
                doc.Close(False)
            excel.Quit()

    def adicionar_imagem_ao_pdf(self, pdf_path, left_image_path, right_image_path, watermark_image_path, image_size_cm=(2, 2)):
        pdf_path = str(pdf_path)
        left_image_path = str(left_image_path)
        right_image_path = str(right_image_path)
        watermark_image_path = str(watermark_image_path)  # Caminho para a imagem da marca d'água

        doc = fitz.open(pdf_path)
        numero_total_paginas = len(doc)  # Obter o número total de páginas
     
        for pagina_number, pagina in enumerate(doc):  # Iterar por todas as páginas
            page_width = pagina.rect.width
            page_height = pagina.rect.height
            texto_contador_paginas = f"- {pagina_number + 1} de {numero_total_paginas} -"  # Formatar o texto do contador

            # Configurar o texto para o contador de páginas
            text_rect = fitz.Rect(0, page_height - 40, page_width, page_height)  # Definir a posição do texto na parte inferior da página
            pagina.insert_textbox(text_rect, texto_contador_paginas, fontsize=11, align=1)  # Inserir o texto do contador
            
            # Inserir marca d'água centralizada em todas as páginas
            wm = fitz.open(watermark_image_path)  # Abrir imagem da marca d'água
            pix = wm[0].get_pixmap()  # Obter pixmap do primeiro documento da imagem
            scale = min(page_width / pix.width, page_height / pix.height) / 1.5  # Escala para reduzir o tamanho da marca d'água
            scaled_width = pix.width * scale
            scaled_height = pix.height * scale
            center_x = (page_width - scaled_width) / 2
            center_y = (page_height - scaled_height) / 2
            watermark_rect = fitz.Rect(center_x, center_y, center_x + scaled_width, center_y + scaled_height)
            
            pagina.insert_image(watermark_rect, filename=watermark_image_path)
            
            # Inserir imagens esquerda e direita apenas na primeira página
            if pagina_number == 0:
                # Calcular o tamanho da imagem em pontos
                image_size_pt = (image_size_cm[0] * 108 / 2.54, image_size_cm[1] * 108 / 2.54)
                
                # Calcular o deslocamento das imagens a partir das bordas em pontos
                offset_left_x_pt = 4 * 72 / 2.54
                offset_right_x_pt = page_width - (4 * 72 / 2.54) - image_size_pt[0]
                offset_y_pt = 0.5 * 72 / 2.54  # 1 cm do topo
                
                # Definir os retângulos onde as imagens serão inseridas
                left_rect = fitz.Rect(offset_left_x_pt, offset_y_pt, offset_left_x_pt + image_size_pt[0], offset_y_pt + image_size_pt[1])
                right_rect = fitz.Rect(offset_right_x_pt, offset_y_pt, offset_right_x_pt + image_size_pt[0], offset_y_pt + image_size_pt[1])
                
                # Inserir as imagens na primeira página
                pagina.insert_image(left_rect, filename=left_image_path)
                pagina.insert_image(right_rect, filename=right_image_path)
            
        # Salvar o documento modificado
        novo_pdf_path = pdf_path.replace('.pdf', '_com_modificacoes.pdf')
        doc.save(novo_pdf_path)
        doc.close()

        # Informar ao usuário sobre o salvamento do novo arquivo
        print(f"PDF modificado salvo como: {novo_pdf_path}")

        # Abrir o PDF automaticamente (Windows)
        try:
            os.startfile(novo_pdf_path)
        except Exception as e:
            print(f"Não foi possível abrir o arquivo PDF automaticamente. Erro: {e}")

    def on_export_pdf(self):
        """
        Exporta os dados para um arquivo PDF e abre o arquivo.
        """
        # Cria um arquivo Excel temporário
        temp_excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_file_path = self.create_excel(temp_excel_file.name)
        
        # Define o caminho para o arquivo PDF de saída
        pdf_file_path = excel_file_path.replace('.xlsx', '.pdf')
        
        # Converte o arquivo Excel em PDF
        self.excel_to_pdf(excel_file_path, pdf_file_path)
        self.adicionar_imagem_ao_pdf(str(pdf_file_path), str(TUCANO_PATH), str(MARINHA_PATH), str(CEIMBRA_BG))
        # Tenta remover o arquivo Excel temporário
        try:
            os.remove(excel_file_path)
        except PermissionError as e:
            print(f"Não foi possível remover o arquivo temporário: {e}. O arquivo pode ainda estar sendo usado.")
        except Exception as e:
            print(f"Erro ao tentar remover o arquivo temporário: {e}")

        # Abre o arquivo PDF gerado
        if sys.platform == "win32":
            os.startfile(pdf_file_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, pdf_file_path])

        print(f"Arquivo PDF exportado e aberto: {pdf_file_path}")

def ajustar_colunas_planilha(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    column_widths = {
        1: 10, 
        2: 10, 
        3: 25, 
        4: 35, 
        5: 0,
        6: 40, 
        7: 10, 
        8: 20, 
        9: 10,
        10: 20, 
        11: 20
    }

    for col_num, width in column_widths.items():
        if width > 0:
            column_letter = openpyxl.utils.get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = width

    workbook.save(file_path)

class ApplicationUI(QMainWindow):
    itemSelected = pyqtSignal(str, str, str)  # Sinal com dois argumentos de string

    NOME_COLUNAS = {
        'mod': 'Mod.',
        'num_pregao': 'N',
        'ano_pregao': 'Ano',
        'nup': 'NUP',
        'objeto': 'Objeto',
        'uasg': 'UASG',
        'orgao_responsavel': 'Órgão Responsável',
        'sigla_om': 'Sigla Órgão',
        'setor_responsavel': 'Demandante',
        'coordenador_planejamento': 'Coordenador',
        'etapa': 'Etapa',
        'pregoeiro': 'Pregoeiro',
    }

    dtypes = {
        'num_pregao': int,
        'ano_pregao': int,
        'mod': str,
        'nup': str,
        'objeto': str,
        'uasg': str,
        'orgao_responsavel': str,
        'sigla_om': str,
        'setor_responsavel': str,
        'coordenador_planejamento': str,
        'etapa': str,
        'pregoeiro': str
    }

    def __init__(self, app, icons_dir, database_dir, lv_final_dir):

        super().__init__()
        self.icons_dir = Path(icons_dir)
        self.database_dir = Path(database_dir)
        self.lv_final_dir = Path(lv_final_dir)
        self.app = app  # Armazenar a instância do App
        
        # Carregar df_uasg uma única vez aqui
        self.df_uasg = pd.read_excel(TABELA_UASG_DIR)
        print("Valores de índices em df_uasg:")
        for index in self.df_uasg.index:
            print(f"Índice: {index}, Valor: {self.df_uasg.loc[index].to_dict()}")
        
        self.columns_treeview = list(self.NOME_COLUNAS.keys())
        self.image_cache = {}
        inicializar_json_do_excel(CONTROLE_PROCESSOS_DIR, PROCESSOS_JSON_PATH)

        # Carregar os dados de licitação no início, removendo a inicialização redundante
        # self.df_licitacao_completo = pd.read_excel(CONTROLE_PROCESSOS_DIR, dtype={'num_pregao': 'Int64'}, converters={'num_pregao': lambda x: self.convert_to_int(x)})
        self.df_licitacao_completo = pd.read_excel(CONTROLE_PROCESSOS_DIR, converters={'num_pregao': lambda x: self.convert_to_int(x)})
        print("Valores de índices em df_licitacao_completo:")
        for index in self.df_licitacao_completo.index:
            print(f"Índice: {index}, Valor: {self.df_licitacao_completo.loc[index].to_dict()}")

        self.image_cache = load_images(self.icons_dir, [
            "plus.png", "save_to_drive.png", "loading.png", "delete.png", "excel.png", "website_menu.png"
        ])
        self.setup_ui()

    def convert_to_int(self, cell_value):
        try:
            return int(cell_value)
        except ValueError:
            return pd.NA  # or some default value or error handling pd.NA  # or a default value like 0 or -1 depending on your requirements

    def _get_image(self, image_file_name):
        # Método para obter imagens do cache ou carregar se necessário
        if image_file_name not in self.image_cache:
            image_path = self.icons_dir / image_file_name
            self.image_cache[image_file_name] = QIcon(str(image_path))  # Usando QIcon para compatibilidade com botões
        return self.image_cache[image_file_name]

    def setup_ui(self):
        self._setup_central_widget()
        self._setup_treeview()  # Configura o QTreeView
        self._adjust_column_widths() 
        self._setup_uasg_delegate()
        self._setup_buttons_layout()
        self.main_layout.addWidget(self.tree_view)
        self._load_data()

    def _setup_central_widget(self):
        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)
        
    def _setup_buttons_layout(self):
        self.buttons_layout = QHBoxLayout()
        self._create_buttons()
        self.main_layout.addLayout(self.buttons_layout)
            
    def _create_buttons(self):
        self.buttons_layout = QHBoxLayout()
        self.button_specs = [
            ("  Adicionar", self.image_cache['plus'], self.on_add_item, "Adiciona um novo item"),
            ("  Salvar", self.image_cache['save_to_drive'], self.on_save_data, "Salva o dataframe em um arquivo excel('.xlsx')"),
            ("  Carregar", self.image_cache['loading'], self.on_load_data, "Carrega o dataframe de um arquivo existente('.xlsx' ou '.odf')"),
            ("  Excluir", self.image_cache['delete'], self.on_delete_item, "Adiciona um novo item"),
            ("  Abrir Planilha Excel", self.image_cache['excel'], self.abrir_planilha_controle, "Abre a planilha de controle"),
            ("    Relatório", self.image_cache['website_menu'], self.on_generate_report, "Gera um relatório dos dados")
        ]

        for text, icon, callback, tooltip in self.button_specs:
            btn = create_button(text=text, icon=icon, callback=callback, tooltip_text=tooltip, parent=self)
            self.buttons_layout.addWidget(btn)  # Adicione o botão ao layout dos botões

    def on_generate_report(self):
        dialog = ReportDialog(self.df_licitacao_completo, self.icons_dir, self)
        dialog.exec()

    def abrir_planilha_controle(self):
        file_path = str(CONTROLE_PROCESSOS_DIR)  # Defina o caminho do arquivo aqui
        try:
            ajustar_colunas_planilha(file_path)
            os.startfile(file_path)
        except Exception as e:
            print(f"Erro ao abrir o arquivo: {e}")

    def _setup_treeview(self):
        # Cria uma nova instância do modelo
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels([self.NOME_COLUNAS[col] for col in self.NOME_COLUNAS])

        # Configurações do QTreeView
        self.tree_view = QTreeView(self)
        self.tree_view.setModel(self.model)
        self.tree_view.setRootIsDecorated(False)
        self.tree_view.setAlternatingRowColors(True)
        self.tree_view.clicked.connect(self._on_item_click)
        self.tree_view.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.model.dataChanged.connect(self._on_item_changed)

        # Adiciona o QTreeView ao layout principal
        self.main_layout.addWidget(self.tree_view)

        # Ajusta as larguras das colunas
        self._adjust_column_widths()

    def _adjust_column_widths(self):
        header = self.tree_view.header()
        header.setStretchLastSection(True)

        # Configura todas as colunas para ajustar-se ao conteúdo
        for column in range(self.model.columnCount()):
            header.setSectionResizeMode(column, QHeaderView.ResizeMode.ResizeToContents)

    def on_context_menu(self, point):
        # Obter o índice do item sob o cursor quando o menu de contexto é solicitado
        index = self.tree_view.indexAt(point)
        
        if index.isValid():
            # Chamar _on_item_click se o índice é válido
            self._on_item_click(index)

            # Criar o menu de contexto
            context_menu = QMenu(self.tree_view)

            # Configurar o estilo do menu de contexto
            context_menu.setStyleSheet("QMenu { font-size: 12pt; }")

            # Adicionar outras ações ao menu
            add_action = context_menu.addAction(QIcon(str(self.icons_dir / "add.png")), "Adicionar")
            edit_action = context_menu.addAction(QIcon(str(self.icons_dir / "engineering.png")), "Editar")
            delete_action = context_menu.addAction(QIcon(str(self.icons_dir / "delete.png")), "Excluir")
            view_action = context_menu.addAction(QIcon(str(self.icons_dir / "search.png")), "Visualizar")

            # Conectar ações a métodos
            add_action.triggered.connect(self.on_add_item)
            edit_action.triggered.connect(self.on_edit_item)
            delete_action.triggered.connect(self.on_delete_item)
            view_action.triggered.connect(self.on_view_item)

            # Executar o menu de contexto na posição do cursor
            context_menu.exec(self.tree_view.viewport().mapToGlobal(point))

    def on_add_item(self):
        # Encontrar o maior número de pregão e adicionar 1
        if not self.model.rowCount():
            novo_num_pregao = 1
        else:
            ultimo_num_pregao = max(int(self.model.item(row, self.columns_treeview.index('num_pregao')).text()) for row in range(self.model.rowCount()))
            novo_num_pregao = ultimo_num_pregao + 1

        # Obter o ano atual
        ano_atual = datetime.datetime.now().year

        # Definir o valor padrão para UASG
        uasg_valor_padrao = "787000"

        # Buscar os dados correspondentes em df_uasg
        uasg_data = self.df_uasg[self.df_uasg['uasg'].astype(str) == uasg_valor_padrao]
        if not uasg_data.empty:
            orgao_responsavel = uasg_data['orgao_responsavel'].iloc[0]
            sigla_om = uasg_data['sigla_om'].iloc[0]
        else:
            orgao_responsavel = "NaN"
            sigla_om = "NaN"
        
        valor_etapa_padrao = "Planejamento"
        mod_padrao = "PE"

        # Criar os valores predefinidos para exibição no QTreeView
        valores_treeview = [
            mod_padrao,
            novo_num_pregao,
            ano_atual,
            f"62055.XXXXXX/{ano_atual}-XX",
            "NaN",  # Objeto
            "787000",
            "NaN",  
        ]

        # Criar uma nova linha no QTreeView com esses valores
        items = [QStandardItem(str(valor)) for valor in valores_treeview]
        self.model.appendRow(items)

        # Criar um dicionário com todos os valores para o DataFrame
        novo_registro = {
            'mod': mod_padrao,
            'num_pregao': novo_num_pregao,
            'ano_pregao': ano_atual,
            'nup': f"62055.XXXXXX/{ano_atual}-XX",
            'objeto': "NaN",
            'uasg': "787000",
            'setor_responsavel': "NaN",
            'orgao_responsavel': orgao_responsavel,  # Colunas adicionais para o DataFrame
            'sigla_om': sigla_om,
            'etapa': valor_etapa_padrao,
            'pregoeiro': "NaN"
        }

        # Verificar se a coluna "etapa" existe no DataFrame, se não, adicioná-la
        if 'etapa' not in self.df_licitacao_completo.columns:
            self.df_licitacao_completo['etapa'] = pd.NA
            
        # Adicionar o novo registro ao DataFrame
        novo_df = pd.DataFrame([novo_registro])
        self.df_licitacao_completo = pd.concat([self.df_licitacao_completo, novo_df], ignore_index=True)

        # Salvar o DataFrame atualizado no arquivo Excel
        save_dataframe_to_excel(self.df_licitacao_completo, CONTROLE_PROCESSOS_DIR)
        sincronizar_json_com_dataframe(self.df_licitacao_completo, PROCESSOS_JSON_PATH)

    def on_edit_item(self):
        # Implementar lógica de edição aqui
        print("Editar item")
    
    def on_save_data(self):
        try:
            # Preencher automaticamente valores NaN na coluna 'etapa' com 'Planejamento'
            self.df_licitacao_completo['etapa'] = self.df_licitacao_completo['etapa'].fillna('Planejamento')

            # Salvar o DataFrame no arquivo Excel
            self.df_licitacao_completo.to_excel(CONTROLE_PROCESSOS_DIR, index=False)
            sincronizar_json_com_dataframe(self.df_licitacao_completo, PROCESSOS_JSON_PATH)

            QMessageBox.information(self, "Sucesso", "Dados salvos com sucesso!")
        except PermissionError:
            QMessageBox.warning(self, "Erro de Permissão", "Não foi possível salvar o arquivo. Por favor, feche o arquivo Excel e tente novamente.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao salvar o arquivo: {str(e)}")

    def on_load_data(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Abrir arquivo", "", "Excel Files (*.xlsx *.xls);;ODF Files (*.odf)")
        if not file_name:
            return 
        try:
            loaded_df = pd.read_excel(file_name, dtype=self.dtypes)
            self.df_licitacao_completo = loaded_df
            self.model.clear()
            self.model.setHorizontalHeaderLabels([self.NOME_COLUNAS[col] for col in self.columns_treeview])

            # Preenche o QTreeView com os dados carregados
            for _, row in self.df_licitacao_completo.iterrows():
                items = [QStandardItem(str(row[col])) for col in self.columns_treeview]
                self.model.appendRow(items)

            # Chama a função para ajustar a largura das colunas
            self._adjust_column_widths()
            sincronizar_json_com_dataframe(self.df_licitacao_completo, PROCESSOS_JSON_PATH)

            QMessageBox.information(self, "Sucesso", "Dados carregados com sucesso do arquivo.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar dados: {e}")

    def on_delete_item(self):
        # Obter o índice do item selecionado
        current_index = self.tree_view.currentIndex()
        if not current_index.isValid():
            QMessageBox.warning(self, "Seleção", "Por favor, selecione um item para excluir.")
            return

        # Obter o número do pregão e o ano do pregão do item selecionado
        row = current_index.row()
        num_pregao = self.model.item(row, self.columns_treeview.index('num_pregao')).text()
        ano_pregao = self.model.item(row, self.columns_treeview.index('ano_pregao')).text()

        # Remover a linha do modelo QTreeView
        self.model.removeRow(row)

        # Atualizar o DataFrame
        self.df_licitacao_completo = self.df_licitacao_completo[
            ~((self.df_licitacao_completo['num_pregao'].astype(str) == num_pregao) &
            (self.df_licitacao_completo['ano_pregao'].astype(str) == ano_pregao))
        ]

        # Salvar o DataFrame atualizado no arquivo Excel
        save_dataframe_to_excel(self.df_licitacao_completo, CONTROLE_PROCESSOS_DIR)
        sincronizar_json_com_dataframe(self.df_licitacao_completo, PROCESSOS_JSON_PATH)

        QMessageBox.information(self, "Sucesso", "Item excluído com sucesso.")

    def on_view_item(self):
        # Implementar lógica de visualização aqui
        print("Visualizar item")

    def _setup_uasg_delegate(self):
        # Configuração do ComboBoxDelegate movida para este método
        uasg_items = [str(item) for item in self.df_uasg['uasg'].tolist()]
        self.uasg_delegate = ComboBoxDelegate(self.tree_view)
        self.uasg_delegate.setItems(uasg_items)
        self.tree_view.setItemDelegateForColumn(self.columns_treeview.index('uasg'), self.uasg_delegate)

        # Carrega os dados no QTreeView
        self._load_data_to_treeview()

    def _load_data_to_treeview(self):
        # Atualiza o modelo com dados atuais do DataFrame
        self.model.clear()  # Limpa o modelo atual
        self.model.setHorizontalHeaderLabels([self.NOME_COLUNAS[col] for col in self.columns_treeview])

        # Preenche o QTreeView com os dados do DataFrame
        for _, row in self.df_licitacao_completo.iterrows():
            items = [QStandardItem(str(row[col])) for col in self.columns_treeview]
            self.model.appendRow(items)

        # Ajusta as larguras das colunas após carregar os dados
        self._adjust_column_widths()

    def _on_item_changed(self, top_left_index, bottom_right_index, roles):
        if Qt.ItemDataRole.EditRole in roles:
            # Salvar a posição atual do scrollbar
            scrollbar = self.tree_view.verticalScrollBar()
            old_scroll_pos = scrollbar.value()

            row = top_left_index.row()
            column = top_left_index.column()
            column_name = self.columns_treeview[column]

            # Obter o valor atualizado
            new_value = str(self.model.itemFromIndex(top_left_index).text())

            # Atualizar o DataFrame se a coluna UASG foi alterada
            if column_name == 'uasg':
                uasg_data = self.df_uasg[self.df_uasg['uasg'].astype(str) == new_value]

                # Se encontrou a UASG correspondente, atualizar as colunas no DataFrame
                if not uasg_data.empty:
                    orgao_responsavel = uasg_data['orgao_responsavel'].iloc[0]
                    sigla_om = uasg_data['sigla_om'].iloc[0]

                    # Atualizar o DataFrame df_licitacao_completo
                    self.df_licitacao_completo.loc[
                        (self.df_licitacao_completo['num_pregao'].astype(str) == self.model.item(row, self.columns_treeview.index('num_pregao')).text()) &
                        (self.df_licitacao_completo['ano_pregao'].astype(str) == self.model.item(row, self.columns_treeview.index('ano_pregao')).text()),
                        ['orgao_responsavel', 'sigla_om']
                    ] = [orgao_responsavel, sigla_om]

            # Obter os valores de identificação únicos (num_pregao e ano_pregao)
            num_pregao = self.model.item(row, self.columns_treeview.index('num_pregao')).text()
            ano_pregao = self.model.item(row, self.columns_treeview.index('ano_pregao')).text()

            # Atualizar o DataFrame para todas as outras colunas
            self.df_licitacao_completo.loc[
                (self.df_licitacao_completo['num_pregao'].astype(str) == num_pregao) &
                (self.df_licitacao_completo['ano_pregao'].astype(str) == ano_pregao),
                column_name
            ] = new_value

            # Salvar o DataFrame atualizado no arquivo Excel
            save_dataframe_to_excel(self.df_licitacao_completo, CONTROLE_PROCESSOS_DIR)
            sincronizar_json_com_dataframe(self.df_licitacao_completo, PROCESSOS_JSON_PATH)

            self._load_data_to_treeview()

            # Restaurar a posição do scrollbar
            scrollbar.setValue(old_scroll_pos)

            # Garantir que a linha editada esteja visível
            self.tree_view.scrollTo(self.model.index(row, 0), QAbstractItemView.ScrollHint.PositionAtCenter)

    def _load_data(self):
        try:
            self.df_licitacao_completo = pd.read_excel(CONTROLE_PROCESSOS_DIR, dtype=self.dtypes)

            # Preencher automaticamente valores NaN na coluna 'etapa' com 'Planejamento'
            self.df_licitacao_completo['etapa'] = self.df_licitacao_completo['etapa'].fillna('Planejamento')

            for _, row in self.df_licitacao_completo.iterrows():
                items = [QStandardItem(str(row[col])) for col in self.NOME_COLUNAS]
                self.model.appendRow(items)
        except Exception as e:
            print(f"Ocorreu um erro ao carregar os dados: {e}")
        self.df_licitacao_exibicao = self.df_licitacao_completo[self.columns_treeview]
        self._populate_treeview()

    def _populate_treeview(self):
        """Populate the treeview with the loaded data."""
        self.model.removeRows(0, self.model.rowCount())
        for index, row in self.df_licitacao_exibicao.iterrows():
            items = [QStandardItem(str(row[col])) for col in self.columns_treeview]
            self.model.appendRow(items)

    def _on_item_click(self, index):
        # Obtenha os valores do item selecionado
        mod = self.model.item(index.row(), self.columns_treeview.index('mod')).text()
        num_pregao = self.model.item(index.row(), self.columns_treeview.index('num_pregao')).text()
        ano_pregao = self.model.item(index.row(), self.columns_treeview.index('ano_pregao')).text()

        print(f"Emitindo sinal para {mod} {num_pregao}/{ano_pregao}")  # Adicione isto para depuração
        self.itemSelected.emit(mod, num_pregao, ano_pregao)

        # Chama o método para processar e salvar o item selecionado
        selected_values = self._get_selected_item_values()
        if selected_values:
            self._process_selected_item(selected_values)

    def _get_selected_item_values(self):
        row = self.tree_view.currentIndex().row()
        if row == -1:
            return []  # Nenhuma linha selecionada

        values = []
        for col in range(self.model.columnCount()):
            item = self.model.item(row, col)
            if item is not None:
                values.append(item.text())
            else:
                values.append("")  # Se não houver item, adicione uma string vazia

        return values

    def _process_selected_item(self, selected_values):
        """Process the selected item."""
        # Recarregar os dados mais recentes do arquivo Excel
        self.df_licitacao_completo = pd.read_excel(CONTROLE_PROCESSOS_DIR, dtype=self.dtypes)

        mod, num_pregao, ano_pregao = selected_values[:3]

        # Filtra o DataFrame completo para encontrar a linha com o num_pregao e ano_pregao correspondentes
        registro_completo = self.df_licitacao_completo[
            (self.df_licitacao_completo['mod'].astype(str).str.strip() == mod) &            
            (self.df_licitacao_completo['num_pregao'].astype(str).str.strip() == num_pregao) &
            (self.df_licitacao_completo['ano_pregao'].astype(str).str.strip() == ano_pregao)
        ]

        if registro_completo.empty:
            # Se nenhum registro for encontrado, retorne False
            return False

        global df_registro_selecionado  # Declare o uso da variável global
        self.itemSelected.emit(mod, num_pregao, ano_pregao)

        df_registro_selecionado = pd.DataFrame(registro_completo)
        df_registro_selecionado.to_csv(ITEM_SELECIONADO_PATH, index=False, encoding='utf-8-sig')
        print(f"Registro salvo em {ITEM_SELECIONADO_PATH}")
        self.app.pregao_selecionado()

        return True

    def run(self):
        """Run the application."""
        self.show()
        self._adjust_column_widths()  

class ComboBoxDelegate(QItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.items = []

    def setItems(self, items):
        self.items = [str(item) for item in items]  # Certifique-se de que todos os itens são strings

    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        editor.addItems(self.items)  # Adiciona itens ao editor
        return editor